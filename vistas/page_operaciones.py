# vistas/page_operaciones.py

import streamlit as st
import pandas as pd
import io
import plotly.express as px
import calendar
from datetime import date, timedelta
import urllib.parse
from controllers.operaciones_controller import OperacionesController
from controllers.venta_controller import VentaController
from controllers.excel_controller import ExcelController
from services.exchange_service import ExchangeService
import importlib

# Asegurar que los controladores se recarguen si hay cambios (Entorno de desarrollo)
try:
    import controllers.excel_controller
    importlib.reload(controllers.excel_controller)
except:
    pass

# NUEVO: Generación de los Excel en memoria cachada (sin widgets de Streamlit)
@st.cache_data(show_spinner=False, ttl=60)
def generate_operational_master_buffers(_controller, id_venta):
    """
    Recopila toda la información de la operación de forma eficiente y genera
    los buffers de Excel correspondientes. Evita widgets de Streamlit para no generar warnings de cache.
    """
    try:
        xl_ctrl = ExcelController()

        # 1. Obtener Datos de la Venta (frescos desde la BD)
        res_v = _controller.client.table('venta').select('*, cliente(nombre, lead(numero_celular, pais_origen)), vendedor(nombre)').eq('id_venta', id_venta).single().execute()
        if not res_v.data:
            return None, None, None
            
        v_raw = res_v.data
        cliente_nest = v_raw.get('cliente', {})
        lead_nest = cliente_nest.get('lead', {}) if isinstance(cliente_nest, dict) else {}
        vendedor_nest = v_raw.get('vendedor', {})

        v_data = {
            "id_venta": v_raw['id_venta'],
            "nombre_cliente": cliente_nest.get('nombre', 'Desconocido') if isinstance(cliente_nest, dict) else 'Desconocido',
            "telefono": lead_nest.get('numero_celular', '---') if isinstance(lead_nest, dict) else '---',
            "tour_nombre": v_raw.get('tour_nombre', 'Sin Tour'),
            "estado_venta": v_raw.get('estado_venta', ''),
            "fecha_venta": v_raw.get('fecha_venta'),
            "fecha_inicio": v_raw.get('fecha_inicio'),
            "fecha_fin": v_raw.get('fecha_fin'),
            "num_pasajeros": v_raw.get('num_pasajeros', 1),
            "vendedor": vendedor_nest.get('nombre', '---') if isinstance(vendedor_nest, dict) else '---',
            "moneda": v_raw.get('moneda', 'USD'),
            "monto_total": v_raw.get('precio_total_cierre', 0),
            "monto_pagado": 0,
            "drive_url": v_raw.get('drive_url'),
            "nro_vuelo_internacional": v_raw.get('nro_vuelo_internacional'),
            "correo_cliente": v_raw.get('correo_cliente'),
            "nombre_contacto_emergencia": v_raw.get('nombre_contacto_emergencia'),
            "telefono_contacto_emergencia": v_raw.get('telefono_contacto_emergencia'),
            "origen": v_raw.get('canal_venta'),
            "telefono_cliente": v_raw.get('telefono_cliente')
        }

        # 2. Calcular Pagos e Información de Depósito
        res_p = _controller.client.table('pago').select('*').eq('id_venta', id_venta).order('fecha_pago', desc=False).execute()
        pagos = res_p.data or []

        ingresos = sum(float(p['monto_pagado'] or 0) for p in pagos if p.get('tipo_pago') != 'REEMBOLSO')
        reembolsos = sum(float(p['monto_pagado'] or 0) for p in pagos if p.get('tipo_pago') == 'REEMBOLSO')

        v_data['monto_pagado'] = ingresos
        v_data['total_reembolsado'] = reembolsos

        if pagos:
            v_data['fecha_primer_deposito'] = pagos[0].get('fecha_pago')
            v_data['metodo_pago_primer'] = pagos[0].get('metodo_pago')
            v_data['monto_primer_deposito'] = pagos[0].get('monto_pagado')
            if len(pagos) > 1:
                v_data['fecha_segundo_deposito'] = pagos[1].get('fecha_pago')
                v_data['metodo_pago_segundo'] = pagos[1].get('metodo_pago')
                v_data['monto_segundo_deposito'] = pagos[1].get('monto_pagado')
            else:
                v_data['fecha_segundo_deposito'] = ""
                v_data['metodo_pago_segundo'] = "---"
                v_data['monto_segundo_deposito'] = 0
        else:
            v_data['fecha_primer_deposito'] = ""
            v_data['metodo_pago_primer'] = ""
            v_data['monto_primer_deposito'] = ""
            v_data['fecha_segundo_deposito'] = ""
            v_data['metodo_pago_segundo'] = ""
            v_data['monto_segundo_deposito'] = ""

        # 3. Obtener Itinerario Logístico
        itinerario = _controller.get_servicios_rango_fechas(date(2000,1,1), date(2100,1,1))
        it_venta = [s for s in itinerario if s['ID Venta'] == id_venta]

        # 4. Obtener Pasajeros
        pasajeros = _controller.pasajero_model.get_by_venta_id(id_venta)

        # 5. Obtener Liquidación Detallada (Costos)
        liquidaciones = _controller.get_liquidaciones_venta(id_venta)

        # 6. Empaquetar y generar Excel
        data_hoja = {
            "venta": v_data,
            "itinerario": it_venta,
            "pasajeros": pasajeros,
            "liquidaciones": liquidaciones
        }

        master_buffer = xl_ctrl.generar_hoja_servicio_maestra_xlsx(data_hoja)
        ficha_buffer = xl_ctrl.generar_ficha_control_grupos_xlsx(data_hoja)
        
        nombre_cliente = v_data.get('nombre_cliente', 'Desconocido')
        
        return master_buffer, ficha_buffer, nombre_cliente
        
    except Exception as e:
        print(f"Error generando data de Hoja de Servicio: {e}")
        return None, None, None

# NUEVO: Renderiza el Botón para el Excel Maestro Operativo (Sin cache).
def render_operational_master_download(controller, id_venta, label="📊 Generar Informe Maestro", key=None):
    """
    Recopila toda la información de la operación utilizando la función cachada y ofrece la descarga de los Excel correspondientes sin warnings de Streamlit.
    """
    try:
        master_buffer, ficha_buffer, nombre_cliente = generate_operational_master_buffers(controller, id_venta)
        
        if master_buffer:
            st.download_button(
                label=label,
                data=master_buffer,
                file_name=f"informe_maestro_{id_venta}_{nombre_cliente.replace(' ', '_')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                help="Expediente completo: Resumen Financiero, Logística, Costos Detallados y Pasajeros.",
                key=key or f"dl_mast_{id_venta}",
                use_container_width=True,
                type="primary"
            )
            
            # --- NUEVO: Botón para Ficha de Control (Replica de Plantilla) ---
            if ficha_buffer:
                st.download_button(
                    label="📋 Ficha de Control (Grupos)",
                    data=ficha_buffer,
                    file_name=f"ficha_control_{id_venta}_{nombre_cliente.replace(' ', '_')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    help="Réplica exacta de la plantilla de control de grupos (Rooming List / Manifiesto).",
                    key=f"dl_ficha_{id_venta}_{key if key else ''}",
                    use_container_width=True
                )
        else:
            st.warning("No se pudo generar el Informe Maestro ni la Ficha de Control en este momento.")
    except Exception as e:
        st.error(f"Error renderizando Hoja de Servicio: {e}")

# Renderiza el Botón para el PDF del Itinerario Simple.
def render_itinerary_simple_download(render):
    if not render:
        st.warning("No hay datos de itinerario para descargar.")
        return

    from controllers.pdf_controller import PDFController
    pdf_ctrl = PDFController()
    
    from controllers.excel_controller import ExcelController
    xl_ctrl = ExcelController()
    
    # Extraer parámetros de enriquecimiento si existen
    nombre_pax = render.get('nombre_pasajero')
    total_pax = render.get('num_pasajeros')
    
    with st.container(border=True):
        st.markdown(f"#### 📄 Resumen de Viaje: {render.get('titulo', 'Sin Título')}")
        st.info("Este documento es una versión simplificada (Ink Saver) ideal para imprimir y para el personal operativo.")
        c1, c2 = st.columns(2)
        
        with c1:
            # Generar el PDF en memoria
            pdf_buffer = pdf_ctrl.generar_itinerario_simple_pdf(render)
            if pdf_buffer:
                st.download_button(
                    label="📥 Bajar Resumen (PDF Simple)",
                    data=pdf_buffer,
                    file_name=f"resumen_viaje_{render.get('titulo', 'itinerario')}.pdf",
                    mime="application/pdf",
                    use_container_width=True
                )
        
        with c2:
            # Generar el Excel en memoria con parámetros forzados
            xlsx_buffer = xl_ctrl.generar_resumen_itinerario_xlsx(render, nombre_cliente=nombre_pax, num_pax=total_pax)
            if xlsx_buffer:
                st.download_button(
                    label="📊 Bajar Resumen (Excel XLSX)",
                    data=xlsx_buffer,
                    file_name=f"resumen_viaje_{render.get('titulo','itin')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
        
        if not pdf_buffer and not xlsx_buffer:
            st.error("No se pudo generar el documento en este momento.")

def render_centro_alertas(controller):
    """
    Componente visual que muestra alertas operativas divididas por urgencia (colores).
    Reglas: 1-2 días Rojo, 3-5 Amarillo, 6-10 Verde.
    Pestaña especial: Machu Picchu (Tickets MINISTERIO).
    """
    with st.expander("🔔 CENTRO DE ALERTAS OPERATIVAS (Urgente)", expanded=True):
        alertas = controller.get_alertas_operativas()
        
        if "error" in alertas and alertas["error"]:
            st.error(f"Error Técnico en Base de Datos: {alertas['error']}")
            with st.expander("Ver Detalle Técnico"):
                st.code(alertas.get('trace', ''))
            return
            
        # --- Cargar comunicados activos ---
        comunicados_activos = []
        rol_actual = str(st.session_state.get('rol_actual', 'VENTAS')).upper()
        col_leido = f"leido_{rol_actual.lower()}"
        
        try:
            from datetime import date as _date
            hoy_iso = _date.today().isoformat()
            res_com = controller.client.table('comunicado').select('*').eq('activo', True).order('fecha_creacion', desc=True).execute()
            
            for c in (res_com.data or []):
                # Validar expiración
                if c.get('fecha_expiracion') and c.get('fecha_expiracion') < hoy_iso:
                    continue
                
                # Validar área de destino
                dest = str(c.get('area_destino', 'TODOS')).upper()
                if dest != 'TODOS' and dest != rol_actual:
                    continue
                
                # Validar si ya fue leído por el área actual
                if c.get(col_leido, False):
                    continue
                
                comunicados_activos.append(c)
        except Exception:
            pass  # Silenciar si la tabla aún no está actualizada en Supabase

        # Pestañas: Comunicados primero, luego operativas
        hay_urgente = any(str(c.get('nivel', '')).upper() == 'URGENTE' for c in comunicados_activos)
        t_com = f"📢 Comunicados ({len(comunicados_activos)})" + (" 🔴" if hay_urgente else "")
        t_mp = f"🏛️ MP Tickets ({len(alertas['machupicchu'])})"
        t_r = f"🔴 Crítico ({len(alertas['rojo'])})"
        t_sa = f"⚠️ Sin Asignar ({len(alertas['sin_asignar'])})"
        t_a = f"🟡 Atención ({len(alertas['amarillo'])})"
        t_v = f"🟢 Preventivo ({len(alertas['verde'])})"
        
        tabs = st.tabs([t_com, t_mp, t_r, t_sa, t_a, t_v])
        
        def mostrar_tabla_alertas(lista_alertas, empty_msg, show_proveedor=True):
            if not lista_alertas:
                st.info(empty_msg)
                return
            
            df = pd.DataFrame(lista_alertas)
            # Ordenar por días
            df = df.sort_values(by="dias")
            
            cols = ["fecha", "servicio", "cliente", "proveedor", "dias"]
            if not show_proveedor:
                cols.remove("proveedor")

            st.dataframe(
                df,
                column_order=cols,
                column_config={
                    "fecha": "📅 Fecha",
                    "servicio": "🚙 Servicio / Tour",
                    "cliente": "👤 Cliente",
                    "proveedor": "🤝 Proveedor",
                    "dias": st.column_config.NumberColumn("⏰ Días Falta", format="%d d")
                },
                use_container_width=True,
                hide_index=True
            )

        # ── PESTAÑA 0: COMUNICADOS ENTRE ÁREAS ──
        with tabs[0]:
            st.markdown("### 📢 Comunicados entre Áreas")
            st.caption("Avisos y alertas enviadas entre las distintas áreas del equipo.")
            
            # Formulario de publicación para todas las áreas
            with st.expander("➕ Publicar Nuevo Comunicado para otra Área"):
                c_t, c_d = st.columns([2, 1])
                titulo_nuevo = c_t.text_input("📝 Título del comunicado", placeholder="Ej: Faltan vouchers de pax X", key="alert_com_titulo")
                
                areas_map = {
                    "Todos": "TODOS",
                    "Ventas": "VENTAS",
                    "Operaciones": "OPERACIONES",
                    "Contabilidad": "CONTABILIDAD",
                    "Gerencia": "GERENCIA"
                }
                dest_nombre = c_d.selectbox("🎯 Dirigido a (Área)", list(areas_map.keys()), key="alert_com_dest")
                dest_code = areas_map[dest_nombre]
                
                c_n, c_u = st.columns([1, 1])
                nivel_nuevo = c_n.selectbox("🔴 Prioridad / Urgencia", ["💡 INFO", "⚠️ AVISO", "🚨 URGENTE"], key="alert_com_nivel")
                
                con_expiracion = c_u.checkbox("Establecer fecha de expiración", key="alert_com_exp_check")
                fecha_exp = None
                if con_expiracion:
                    from datetime import date as _date, timedelta
                    fecha_exp = st.date_input("Expira el", value=_date.today() + timedelta(days=7), key="alert_com_fecha_exp")
                
                mensaje_nuevo = st.text_area("💬 Mensaje o Detalle", placeholder="Escribe aquí los detalles del aviso...", height=80, key="alert_com_mensaje")
                
                if st.button("🚀 Publicar Comunicado", type="primary", use_container_width=True, key="alert_btn_publish"):
                    if not titulo_nuevo.strip() or not mensaje_nuevo.strip():
                        st.warning("⚠️ El Título y el Mensaje son obligatorios.")
                    else:
                        nivel_code = nivel_nuevo.split(" ")[-1]
                        payload = {
                            'titulo': titulo_nuevo.strip(),
                            'mensaje': mensaje_nuevo.strip(),
                            'nivel': nivel_code,
                            'autor_area': rol_actual,
                            'area_destino': dest_code,
                            'activo': True,
                            'fecha_expiracion': fecha_exp.isoformat() if fecha_exp else None
                        }
                        try:
                            controller.client.table('comunicado').insert(payload).execute()
                            st.success("✅ Comunicado enviado exitosamente a la(s) área(s) destino.")
                            st.rerun()
                        except Exception as ex:
                            st.error(f"Error al publicar comunicado: {ex}")
            
            st.markdown("---")
            
            if not comunicados_activos:
                st.info("No hay comunicados pendientes para tu área en este momento.")
            else:
                _nivel_cfg = {
                    'URGENTE': ('🚨', 'error'),
                    'AVISO':   ('⚠️', 'warning'),
                    'INFO':    ('💡', 'info'),
                }
                for c in comunicados_activos:
                    nivel = str(c.get('nivel', 'INFO')).upper()
                    emoji, fn_name = _nivel_cfg.get(nivel, ('📌', 'info'))
                    fecha_raw = str(c.get('fecha_creacion', ''))[:10]
                    autor_area = c.get('autor_area', 'GERENCIA')
                    area_destino = c.get('area_destino', 'TODOS')
                    
                    with st.container(border=True):
                        # Encabezado del nivel
                        if fn_name == 'error':
                            st.error(f"{emoji} **{c.get('titulo', 'Sin título')}** (URGENTE)")
                        elif fn_name == 'warning':
                            st.warning(f"{emoji} **{c.get('titulo', 'Sin título')}** (AVISO)")
                        else:
                            st.info(f"{emoji} **{c.get('titulo', 'Sin título')}**")
                            
                        st.write(c.get('mensaje', ''))
                        st.caption(f"De: **{autor_area}** ➔ Para: **{area_destino}** &nbsp;|&nbsp; 📅 {fecha_raw}")
                        
                        # Botón para marcar Check
                        if st.button("☑️ Confirmar recepción / Marcar check (Desaparece)", key=f"chk_read_{c['id']}", use_container_width=True):
                            try:
                                controller.client.table('comunicado').update({col_leido: True}).eq('id', c['id']).execute()
                                st.success("¡Marcado como leído para tu área!")
                                st.rerun()
                            except Exception as ex:
                                st.error(f"Error al marcar check: {ex}")

        with tabs[1]:
            st.markdown("### 🏛️ Tickets Machu Picchu (MINISTERIO)")
            st.caption("Todos los ingresos pendientes asignados al proveedor del estado.")
            mostrar_tabla_alertas(alertas['machupicchu'], "No hay tickets de MP pendientes.")
            
        with tabs[2]:
            st.markdown("### 🔴 Alertas Críticas (0 a 2 días)")
            st.caption("Servicios inminentes que no han sido marcados como 'Terminado'.")
            mostrar_tabla_alertas(alertas['rojo'], "No hay alertas críticas pendientes.")

        with tabs[3]:
            st.markdown("### ⚠️ Servicios Sin Asignar (Riesgo Alto)")
            st.error("Estos servicios están en el itinerario pero NO tienen costos ni proveedores registrados. Las alertas de colores NO funcionarán para estos casos si no se completan.")
            mostrar_tabla_alertas(alertas['sin_asignar'], "Todos los servicios dentro de los próximos 10 días tienen asignaciones.")
            
        with tabs[4]:
            st.markdown("### 🟡 Alertas de Atención (3 a 5 días)")
            st.caption("Servicios próximos que requieren verificación operativa.")
            mostrar_tabla_alertas(alertas['amarillo'], "No hay alertas de atención pendientes.")
            
        with tabs[5]:
            st.markdown("### 🟢 Alertas Preventivas (6 a 10 días)")
            st.caption("Servicios programados para la próxima semana.")
            mostrar_tabla_alertas(alertas['verde'], "No hay alertas preventivas pendientes.")

# Dashboard 2: Tablero con vistas Duplicadas (Mensual/Semanal).
def dashboard_tablero_diario(controller):
    """Dashboard 2: Tablero con vistas Duplicadas (Mensual/Semanal)."""
    st.subheader("2️⃣ Tablero de Planificación Logística", divider='green')

    # --- LEYENDA DE COLORES ---
    st.markdown(
        """
        <div style='display:flex; flex-wrap:wrap; gap:8px; margin-bottom:10px;'>
            <span style='background:#1a1a1a;border:1px solid #333;border-radius:20px;padding:4px 12px;font-size:12px;'>🔴 Registrado (sin checks)</span>
            <span style='background:#1a1a1a;border:1px solid #333;border-radius:20px;padding:4px 12px;font-size:12px;'>🟡 En progreso (algún check)</span>
            <span style='background:#1a1a1a;border:1px solid #333;border-radius:20px;padding:4px 12px;font-size:12px;'>🟢 Confirmado (todos los checks)</span>
            <span style='background:#1a1a1a;border:1px solid #333;border-radius:20px;padding:4px 12px;font-size:12px;'>🔵 Contratando (algún contrato)</span>
            <span style='background:#1a1a1a;border:1px solid #333;border-radius:20px;padding:4px 12px;font-size:12px;'>🟠 Contratado (todos los contratos)</span>
            <span style='background:#1a1a1a;border:1px solid #333;border-radius:20px;padding:4px 12px;font-size:12px;'>🟣 Cancelado</span>
            <span style='background:#1a1a1a;border:1px solid #333;border-radius:20px;padding:4px 12px;font-size:12px;'>⬜ Aprobado por Gerencia</span>
        </div>
        """,
        unsafe_allow_html=True
    )

    if 'cal_current_date' not in st.session_state:
        st.session_state['cal_current_date'] = date.today()
    if 'cal_selected_date' not in st.session_state:
        st.session_state['cal_selected_date'] = date.today()
    if 'view_mode' not in st.session_state:
        st.session_state['view_mode'] = "Mensual"

    v_mode = st.radio("Filtro de Vista:", ["Mensual", "Semanal"], 
                      index=0 if st.session_state['view_mode'] == "Mensual" else 1, horizontal=True)
    st.session_state['view_mode'] = v_mode

    current_date = st.session_state['cal_current_date']
    year, month = current_date.year, current_date.month
    nombres_meses = ["", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", 
                     "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]

    if st.session_state['view_mode'] == "Mensual":
        # --- MES ---
        c1, c2, c3 = st.columns([1, 2, 1])
        with c1:
            if st.button("◀ Mes Ant"):
                m, y = (12, year-1) if month == 1 else (month-1, year)
                st.session_state['cal_current_date'] = date(y, m, 1)
                st.rerun()
        with c2:
            st.markdown(f"<h3 style='text-align:center;'>{nombres_meses[month]} {year}</h3>", unsafe_allow_html=True)
        with c3:
            if st.button("Mes Sig ▶"):
                m, y = (1, year+1) if month == 12 else (month+1, year)
                st.session_state['cal_current_date'] = date(y, m, 1)
                st.rerun()

        st.markdown("---")
        cal_grid = calendar.monthcalendar(year, month)
        # Obtener colores reales de cada día del mes
        fechas_colores = controller.get_fechas_con_colores(year, month)
        
        cols = st.columns(7)
        headers = ['Lun', 'Mar', 'Mié', 'Jue', 'Vie', 'Sáb', 'Dom']
        for i, h in enumerate(headers): cols[i].markdown(f"<center><b>{h}</b></center>", unsafe_allow_html=True)
            
        for week in cal_grid:
            cols = st.columns(7)
            for i, day in enumerate(week):
                if day != 0:
                    d_obj = date(year, month, day)
                    sel = (d_obj == st.session_state['cal_selected_date'])
                    color_emoji = fechas_colores.get(d_obj, '')  # '' si no hay servicios ese día
                    lbl = f"{day}{(' ' + color_emoji) if color_emoji else ''}"
                    if d_obj == date.today(): lbl += "\n(Hoy)"
                    if cols[i].button(lbl, key=f"d_{d_obj}", use_container_width=True, type="primary" if sel else "secondary"):
                        st.session_state['cal_selected_date'] = d_obj
                        st.rerun()

    else:
        # --- SEMANA ---
        d_sel = st.session_state['cal_selected_date']
        lunes = d_sel - timedelta(days=d_sel.weekday())
        domingo = lunes + timedelta(days=6)
        
        c1, c2, c3 = st.columns([1, 2, 1])
        with c1:
            if st.button("◀ Semana Ant"):
                st.session_state['cal_selected_date'] -= timedelta(days=7)
                st.rerun()
        with c2:
            st.markdown(f"<h3 style='text-align:center;'>Semana de {lunes.day} {nombres_meses[lunes.month]}</h3>", unsafe_allow_html=True)
        with c3:
            if st.button("Semana Sig ▶"):
                st.session_state['cal_selected_date'] += timedelta(days=7)
                st.rerun()
        
        st.markdown("---")
        servicios_w = controller.get_servicios_rango_fechas(lunes, domingo)
        cols_w = st.columns(7)
        headers_w = ['Lun', 'Mar', 'Mié', 'Jue', 'Vie', 'Sáb', 'Dom']
        
        for i in range(7):
            f_dia = lunes + timedelta(days=i)
            with cols_w[i]:
                estilo = f"background:{'#1E88E5' if f_dia==date.today() else '#444'}; padding:5px; border-radius:5px; text-align:center; margin-bottom:5px;"
                st.markdown(f"<div style='{estilo}'><small>{headers_w[i]}</small><br><b>{f_dia.day}</b></div>", unsafe_allow_html=True)
                
                s_dia = [s for s in servicios_w if s['Fecha'] == f_dia.isoformat()]
                if not s_dia:
                    st.markdown("<p style='text-align:center; color:gray; font-size:10px;'>Vacío</p>", unsafe_allow_html=True)
                else:
                    for s in s_dia:
                        with st.container(border=True):
                            # Título con indicador de endoso
                            titulo_serv = f"🤝 {s['Servicio']}" if s.get('Endoso?') else s['Servicio']
                            st.markdown(f"<p style='font-size:11px; margin:0;'><b>{titulo_serv}</b></p>", unsafe_allow_html=True)
                            st.markdown(f"<p style='font-size:9px; margin:0; color:#aaa;'>{s['Cliente']} ({s['Pax']} Pax)</p>", unsafe_allow_html=True)
                            
                            # Responsable (Guía o Agencia)
                            responsable = f"🏢 {s['Agencia Endoso']}" if s.get('Endoso?') else f"👮 {s['Guía']}"
                            st.markdown(f"<p style='font-size:9px; margin:0;'>{responsable}</p>", unsafe_allow_html=True)
                
                if st.button("Ver", key=f"v_{f_dia}", use_container_width=True):
                    st.session_state['cal_selected_date'] = f_dia
                    st.rerun()

    # --- DETALLE ---
    st.markdown("---")
    f_p = st.session_state['cal_selected_date']
    st.markdown(f"### 📋 Detalle: {f_p.day} de {nombres_meses[f_p.month]} de {f_p.year}")
    
    servicios = controller.get_servicios_por_fecha(f_p)
    if not servicios:
        st.info("Sin operaciones este día.")
    else:
        pax_val = 0
        try: pax_val = sum(int(s.get('Pax') or 0) for s in servicios)
        except: pass
        # --- TABLAR INTERACTIVA DE SERVICIOS (REDISEÑO) ---
        c_head = st.columns([0.8, 2.5, 0.6, 2, 1.2])
        c_head[0].markdown("**Hora**")
        c_head[1].markdown("**Servicio**")
        c_head[2].markdown("**Pax**")
        c_head[3].markdown("**Cliente**")
        c_head[4].markdown("**📄 Info**")
        st.markdown("<hr style='margin:0; border:0.5px solid #555;'>", unsafe_allow_html=True)

        for i, s in enumerate(servicios):
            with st.container():
                c_row = st.columns([0.8, 2.5, 0.6, 2, 1.2])
                c_row[0].write(s.get('Hora', '---'))
                
                # Nombre del servicio con indicador de endoso
                serv_name = f"🤝 {s['Servicio']}" if s.get('Endoso?') else s['Servicio']
                c_row[1].write(f"**{serv_name}**")
                
                c_row[2].write(f"**{s.get('Pax', 1)}**")

                # Mostrar cliente con badge de color semáforo
                id_v = s.get('ID Venta')
                color_badge = controller.get_color_venta(id_v) if id_v else ''
                c_row[3].write(f"{color_badge} {s.get('Cliente', '---')}")
                
                with c_row[4]:
                    if id_v:
                        render_operational_master_download(controller, id_v, label="📁 Maestro", key=f"dl_cal_{id_v}_{i}")
                    else:
                        st.caption("Sin Venta")
                st.markdown("<hr style='margin:0; border:0.1px solid #333;'>", unsafe_allow_html=True)

        st.info("💡 Haz clic en '📁 Maestro' para descargar el expediente completo del pasajero.")

def generar_mensaje_whatsapp(data):
    """Genera un link de WhatsApp con el mensaje formateado."""
    texto = f"Hola, soy de la Agencia. Aquí el detalle de tu servicio:\n\n"
    texto += f"Servicio: {data['Servicio']}\n"
    texto += f"Cliente: {data['Cliente']}\n"
    texto += f"Fecha: {data['Fecha']}\n"
    texto += f"Guía: {data['Guía']}"
    
    url = f"https://wa.me/?text={urllib.parse.quote(texto)}"
    return url


def registro_ventas_proveedores(supabase_client):
    from controllers.itinerario_digital_controller import ItinerarioDigitalController
    from controllers.lead_controller import LeadController
    import json

    venta_controller = VentaController(supabase_client)
    it_controller = ItinerarioDigitalController(supabase_client)
    lead_controller = LeadController(supabase_client)
                                                                                                                                                            
    st.subheader("🤝 Registro de Venta B2B (Agencias & Partners)")

    # ═══════════════════════════════════════════════════════════════
    # 1️⃣ Buscador Inteligente de Lead y Selector de Itinerario
    # ═══════════════════════════════════════════════════════════════
    search_query = st.text_input("🔍 Buscar Pasajero (Nombre o Celular)", placeholder="Escriba para filtrar...", key="search_lead_b2b").strip().lower()
    
    leads = lead_controller.obtener_todos_leads()
    lead_opt = ["--- Selecciona un Lead (Obligatorio) ---"]
    lead_map = {}
    
    if leads:                                                                                                                                                           
        if search_query:
            filtered_leads = [
                l for l in leads 
                if search_query in str(l.get('nombre_pasajero', '')).lower() or 
                   search_query in str(l.get('numero_celular', '')).lower()
            ]
            st.caption(f"✨ Se encontraron {len(filtered_leads)} coincidencias.")
        else:
            filtered_leads = leads[:20]
            st.caption("💡 Mostrando los 20 más recientes. Use el buscador para ver otros.")

        for l in filtered_leads:
            lbl = f"{l['numero_celular']} - {l.get('nombre_pasajero') or 'Sin Nombre'}"
            lead_opt.append(lbl)
            lead_map[lbl] = l

    lead_sel = st.selectbox("🎯 Vincular con un Lead existente", lead_opt, help="Busque por nombre o celular arriba", key="lead_sel_b2b")
    lead_data = lead_map.get(lead_sel)

    id_lead_seleccionado = lead_data.get('id_lead') if lead_data else None
    
    if id_lead_seleccionado:
        itinerarios_recuperados = it_controller.listar_itinerarios_lead(id_lead_seleccionado)
    else:
        itinerarios_recuperados = it_controller.obtener_todos_recientes(limit=30)
    
    opciones_itinerario = ["--- Sin Itinerario ---"]
    mapa_itinerarios = {}
    
    if itinerarios_recuperados:
        itinerarios_recuperados.sort(key=lambda x: x.get('fecha_generacion', ''))
        conteos = {}

        for it in itinerarios_recuperados:
            id_lead_from_itinerario = it.get('id_lead')
            render_data = it.get('datos_render', {})
            if isinstance(render_data, str):
                import json
                try: render_data = json.loads(render_data)
                except: render_data = {}

            # Ya no filtramos por B2B o B2C, mostramos todos los itinerarios del pasajero
            # ya que la pestaña define el tipo de venta final.

            titulo = render_data.get('titulo', '')
            if not titulo:
                title_1 = render_data.get('title_1', '')
                title_2 = render_data.get('title_2', '')
                titulo = f"{title_1} {title_2}".strip() or 'Sin título'
            
            fecha = it.get('fecha_generacion', '')[:10] if it.get('fecha_generacion') else 'Sin fecha'
            
            celular = it.get('lead', {}).get('numero_celular', '') if it.get('lead') else lead_data.get('numero_celular', '') if lead_data else ''
            cel_label = f"📱 {celular} | " if celular else ""
            
            base_label = f"{cel_label}{titulo} ({fecha})"
            
            conteos[base_label] = conteos.get(base_label, 0) + 1
            ver = conteos[base_label]
            
            label_final = f"{base_label} - V{ver}"
            opciones_itinerario.append(label_final)
            mapa_itinerarios[label_final] = it
    


    itinerario_seleccionado = st.selectbox(
        "✨ Seleccionar Itinerario Visual (Diseño Cloud)", 
        opciones_itinerario,
        help="Seleccione el diseño que corresponde a esta venta B2B",
        key="itin_sel_b2b"
    )

    # ═══════════════════════════════════════════════════════════════
    # 2️⃣ AUTO-COMPLETADO Y DATOS SUGERIDOS
    # ═══════════════════════════════════════════════════════════════
    id_itinerario_dig = None
    def_pax = ""
    def_tour = ""
    def_cel = ""
    def_precio_total = 0.0
    def_moneda = 'USD'
    def_f_inicio = date.today()
    def_f_fin = date.today()
    def_cant_pax = 1

    if itinerario_seleccionado != "--- Sin Itinerario ---":
        it_data = mapa_itinerarios.get(itinerario_seleccionado)
        if it_data:
            id_itinerario_dig = it_data.get('id_itinerario_digital')
            id_lead_from_itinerario = it_data.get('id_lead')
            render = it_data.get('datos_render', {})
            if isinstance(render, str):
                import json
                try: render = json.loads(render)
                except: render = {}

            def_pax = it_data.get('nombre_pasajero_itinerario', '') or render.get('pasajero', '') or def_pax
            def_tour = render.get('titulo', '') or f"{render.get('title_1', '')} {render.get('title_2', '')}".strip()
            
            # Extraer celular del lead o itinerario
            cel_cloud = ''
            if it_data.get('lead') and isinstance(it_data['lead'], dict):
                cel_cloud = it_data['lead'].get('numero_celular', '')
            if not cel_cloud and lead_data:
                cel_cloud = lead_data.get('numero_celular', '')
            def_cel = cel_cloud or def_cel
            
            # --- 📅 CÁLCULO AUTOMÁTICO DE FECHAS (ROBUSTO) ---
            f_viaje = render.get('fecha_viaje') or render.get('fecha_inicio') or render.get('fechaViaje') or render.get('fecha')
            if not f_viaje and render.get('control_interno'):
                ci = render.get('control_interno', {})
                f_viaje = ci.get('fecha_inicio') or ci.get('fecha_llegada') or ci.get('fecha_viaje')
            if not f_viaje:
                dias = render.get('itinerario') or render.get('days') or render.get('itinerario_detalles')
                if dias and isinstance(dias, list) and len(dias) > 0 and isinstance(dias[0], dict):
                    f_viaje = dias[0].get('fecha')
                
            if f_viaje:
                try: 
                    f_clean = str(f_viaje).replace(" ", "").strip()
                    if '/' in f_clean:
                        from datetime import datetime
                        try:
                            def_f_inicio = datetime.strptime(f_clean, "%d/%m/%Y").date()
                        except ValueError:
                            try: def_f_inicio = datetime.strptime(f_clean, "%Y/%m/%d").date()
                            except ValueError: def_f_inicio = datetime.strptime(f_clean, "%m/%d/%Y").date()
                    elif '-' in f_clean:
                        def_f_inicio = date.fromisoformat(f_clean[:10])
                except: pass
            
            def_f_fin = def_f_inicio
            duracion_raw = render.get('duracion')
            if duracion_raw and isinstance(duracion_raw, str) and 'D' in duracion_raw.upper():
                try:
                    num_dias_str = ''.join(filter(str.isdigit, duracion_raw.split('D')[0]))
                    if num_dias_str: def_f_fin = def_f_inicio + timedelta(days=int(num_dias_str) - 1)
                except: pass
            
            # --- CÁLCULO DE PAX ROBUSTO ---
            if render.get('control_interno'):
                def_cant_pax = int(render['control_interno'].get('total_pasajeros') or render['control_interno'].get('total_pax') or 1)
            elif render.get('detalle_ingresos'):
                def_cant_pax = sum(int(d.get('cantidad', 0)) for d in render['detalle_ingresos'])
            else:
                def_cant_pax = int(render.get('cantidad_pax') or render.get('pax_count') or 1)

            # Sincronizar con el widget del formulario
            st.session_state['b2b_pax_v2'] = def_cant_pax

            # --- INICIALIZACIÓN DE VARIABLES PARA AUTO-COMPLETADO ---
            items_extraidos = []
            tipos_vistos = set()
            ci = render.get('control_interno', {})
            tc_itin = ci.get('tipo_cambio_aplicado', 3.8)
            try: tc_itin = float(tc_itin or 3.8)
            except: tc_itin = 3.8

            # 1. PRIORIDAD: Precios Cierre (Itinerario Automático)
            precios_cierre = render.get('precios_cierre', [])
            if precios_cierre and isinstance(precios_cierre, list):
                for pc in precios_cierre:
                    label = pc.get('label', '').upper()
                    m_raw = pc.get('monto_total', 0) or 0
                    try: monto_total = float(str(m_raw).replace(',', ''))
                    except: monto_total = 0.0
                    simbolo = pc.get('simbolo', '$')
                    
                    t_code = 'NACIONAL' if 'NACIONAL' in label else ('EXTRANJERO' if 'EXTRANJERO' in label else 'CAN')
                    es_usd = (simbolo == '$')
                    
                    # Conteo de pax desde desglose_pasajeros
                    pax_info = ci.get('desglose_pasajeros', {}).get(t_code.lower(), {})
                    c_f = sum(int(v or 0) for v in pax_info.values()) if isinstance(pax_info, dict) else 0
                    if not c_f: c_f = 1 # Fallback a 1 pax para no romper división
                    
                    p_f_soles = monto_total * tc_itin if es_usd else monto_total
                    items_extraidos.append({
                        "descripcion": f"Pax {t_code.capitalize()} (Auto-Total)", 
                        "cantidad": int(c_f), 
                        "precio_unitario": p_f_soles / int(c_f), 
                        "tipo": t_code, 
                        "p_raw": monto_total / int(c_f),
                        "moneda": "USD" if es_usd else "PEN"
                    })
                    tipos_vistos.add(t_code)

            # 2. SEGUNDA OPCIÓN: Precios Estructurados (Manual Nuevo o Auto-Compacto)
            precios_meta = render.get('precios', {})
            if not items_extraidos and precios_meta:
                # Mapeo de claves posibles
                mapeo_p = [('NACIONAL', ['nacional', 'nac']), ('EXTRANJERO', ['extranjero', 'ext']), ('CAN', ['can'])]
                for t_code, keys in mapeo_p:
                    p_data = None
                    for k in keys:
                        p_data = precios_meta.get(k)
                        if p_data is not None: break
                    
                    if p_data is not None:
                        # Extraer monto
                        try:
                            if isinstance(p_data, dict): 
                                m_raw = p_data.get('total') or p_data.get('monto') or 0
                                p_raw = float(str(m_raw).replace(',', ''))
                            else: 
                                p_raw = float(str(p_data).replace(',', ''))
                        except:
                            p_raw = 0.0
                        
                        if p_raw > 0:
                            # Moneda
                            moneda_key = f"moneda_{t_code.lower()}"
                            moneda_fix = precios_meta.get(moneda_key)
                            es_usd = (moneda_fix == "USD") if moneda_fix else (t_code in ['EXTRANJERO', 'CAN'])
                            
                            # Conteo
                            c_f = int(render.get(f'num_pax_{t_code.lower()[0:3]}', 1) or 1)
                            
                            p_f_soles = p_raw * tc_itin if es_usd else p_raw
                            items_extraidos.append({
                                "descripcion": f"Pax {t_code.capitalize()} (Precio)", 
                                "cantidad": c_f, "precio_unitario": p_f_soles, "tipo": t_code, "p_raw": p_raw,
                                "moneda": "USD" if es_usd else "PEN"
                            })
                            tipos_vistos.add(t_code)

            # 3. FALLBACK LEGACY: Raíz (num_pax_nac, etc.)
            if not items_extraidos:
                fallbacks = [
                    ('NACIONAL', ['num_pax_nac', 'pax_nac', 'num_pax_nacional'], ['precio_nacional', 'p_nac']),
                    ('EXTRANJERO', ['num_pax_ext', 'pax_ext', 'num_pax_extranjero'], ['precio_extranjero', 'p_ext']),
                    ('CAN', ['num_pax_can', 'pax_can'], ['precio_can', 'p_can'])
                ]
                for t_code, c_keys, p_keys in fallbacks:
                    if t_code not in tipos_vistos:
                        c_f = 0
                        for ck in c_keys:
                            c_f = int(render.get(ck, 0) or 0)
                            if c_f > 0: break
                        if c_f > 0:
                            p_f_raw = 0.0
                            for pk in p_keys:
                                val_pk = render.get(pk, 0) or 0
                                try: p_f_raw = float(str(val_pk).replace(',', ''))
                                except: p_f_raw = 0.0
                                if p_f_raw > 0: break
                            p_f_soles = p_f_raw * tc_itin if t_code in ['EXTRANJERO', 'CAN'] else p_f_raw
                            items_extraidos.append({
                                "descripcion": f"Pax {t_code.capitalize()} (Legacy)", "cantidad": c_f, "precio_unitario": p_f_soles, "tipo": t_code, "p_raw": p_f_raw,
                                "moneda": "USD" if t_code in ['EXTRANJERO', 'CAN'] else "PEN"
                            })
                pax_gen = render.get('cantidad_pax') or render.get('pax_count') or render.get('num_pax') or 0
                if not pax_gen and ci: 
                    pax_gen = ci.get('total_pasajeros') or ci.get('total_pax') or 0
                
                if pax_gen:
                    p_sug_raw = render.get('total_final_calculado') or render.get('precio_cierre') or 0
                    try: p_sug_val = float(str(p_sug_raw).replace(',', ''))
                    except: p_sug_val = 0.0
                    items_extraidos.append({
                        "descripcion": "Pax (Itinerario)", 
                        "cantidad": int(pax_gen), 
                        "precio_unitario": p_sug_val / int(pax_gen) if int(pax_gen) > 0 else 0,
                        "tipo": "NACIONAL", 
                        "p_raw": p_sug_val
                    })

            # Cálculo de Total Final en Soles
            total_soles = sum(it['cantidad'] * it['precio_unitario'] for it in items_extraidos)

            if id_itinerario_dig:
                # Actualizar siempre los items en sesión B2B
                st.session_state[f"b2b_items_{id_itinerario_dig}"] = items_extraidos

                if id_itinerario_dig != st.session_state.get('b2b_last_itin_v2'):
                    st.session_state['b2b_m_total'] = total_soles
                    st.session_state['b2b_moneda_auto'] = 'PEN'
                    st.session_state['b2b_last_itin_v2'] = id_itinerario_dig
                    st.success(f"✅ Itinerario cargado: **{def_tour}** (Total calculado: S/ {total_soles:,.2f})")
                else:
                    st.success(f"✅ Itinerario cargado: **{def_tour}**")

    # ═══════════════════════════════════════════════════════════════
    # 4️⃣ BALANCE INTERACTIVO (Igual que Ventas Directas)
    # ═══════════════════════════════════════════════════════════════
    st.markdown("### 💰 Detalles de Pago B2B")
    c_p0, c_p1, c_p2, c_p3 = st.columns([1, 1.5, 1.5, 1.5])
    
    monedas_list = ["USD", "PEN"]
    m_auto = st.session_state.get('b2b_moneda_auto', 'USD')
    
    # --- MOSTRAR SUB-TOTALES POR NACIONALIDAD B2B (PEDIDO USUARIO) ---
    if id_itinerario_dig:
        items_ref_b2b = st.session_state.get(f"b2b_items_{id_itinerario_dig}", [])
        if items_ref_b2b:
            # Ahora filtramos por la MONEDA real del item, no solo el tipo
            sub_soles = sum(it['cantidad'] * it['p_raw'] for it in items_ref_b2b if it.get('moneda') == 'PEN')
            sub_dolares = sum(it['cantidad'] * it['p_raw'] for it in items_ref_b2b if it.get('moneda') == 'USD')
            
            st.markdown(f"📊 **SUB-TOTALES POR MONEDA:** Soles: **S/ {sub_soles:,.2f}** | Dólares: **$ {sub_dolares:,.2f}**")

    idx_m = monedas_list.index(m_auto) if m_auto in monedas_list else 0
    moneda_sel = c_p0.selectbox("Moneda", monedas_list, index=idx_m, key="b2b_final_moneda")
    
    # TC: Tipo de Cambio "Foto"
    default_tc = ExchangeService.get_current_tc()
    tipo_cambio = c_p1.number_input("TC (Foto)", min_value=0.0, value=default_tc, format="%.3f", key="b2b_tc")

    # --- RECÁLCULO DINÁMICO B2B: Usar el TC del usuario para actualizar el total ---
    if id_itinerario_dig and tipo_cambio > 0:
        items_recalc_b2b = st.session_state.get(f"b2b_items_{id_itinerario_dig}", [])
        if items_recalc_b2b:
            # Calculamos el total base siempre en SOLES primero
            total_base_soles = 0.0
            for it in items_recalc_b2b:
                if it.get('moneda') == "USD":
                    total_base_soles += it['cantidad'] * it['p_raw'] * tipo_cambio
                else:
                    total_base_soles += it['cantidad'] * it['p_raw']
            
            # Ahora, convertimos ese total base a la moneda seleccionada para el input
            if moneda_sel == "USD":
                st.session_state['b2b_m_total'] = round(total_base_soles / tipo_cambio, 2)
            else:
                st.session_state['b2b_m_total'] = round(total_base_soles, 2)

    if 'b2b_m_total' not in st.session_state: st.session_state['b2b_m_total'] = 0.0
    if 'b2b_m_pago' not in st.session_state: st.session_state['b2b_m_pago'] = 0.0
    
    monto_total = c_p2.number_input(f"Monto Total ({moneda_sel})", min_value=0.0, format="%.2f", key="b2b_m_total", disabled=(id_itinerario_dig is not None))
    monto_pagado = c_p3.number_input(f"Adelanto Agencia ({moneda_sel})", min_value=0.0, format="%.2f", key="b2b_m_pago")
    
    # --- MOSTRAR CONVERSIÓN EN TIEMPO REAL (B2B) ---
    if tipo_cambio > 0:
        if moneda_sel == "USD":
            c_p2.caption(f"🛡️ Equiv: **S/ {monto_total * tipo_cambio:,.2f}**")
            c_p3.caption(f"🛡️ Equiv: **S/ {monto_pagado * tipo_cambio:,.2f}**")
        else:
            c_p2.caption(f"🛡️ Equiv: **$ {monto_total / tipo_cambio:,.2f}**")
            c_p3.caption(f"🛡️ Equiv: **$ {monto_pagado / tipo_cambio:,.2f}**")

    saldo = monto_total - monto_pagado
    if monto_total > 0:
        if saldo <= 0.01: 
            st.success(f"✅ **VENTA B2B SALDADA**")
        else:
            # Mostrar saldo bilingüe
            if moneda_sel == "USD":
                info_s = f"⏳ **SALDO PENDIENTE AGENCIA: ${saldo:,.2f}** (S/ {saldo * tipo_cambio:,.2f})"
            else:
                info_s = f"⏳ **SALDO PENDIENTE AGENCIA: S/ {saldo:,.2f}** (${saldo / tipo_cambio:,.2f})"
            st.warning(info_s)

    # ═══════════════════════════════════════════════════════════════
    # 5️⃣ FORMULARIO DE REGISTRO
    # ═══════════════════════════════════════════════════════════════
    with st.form("form_b2b_redesigned"):
        col1, col2 = st.columns(2)
        
        # Agencia / Proveedor (Mandatorio)
        agencias = venta_controller.obtener_agencias_aliadas()
        nombres_ag = [a['nombre'] for a in agencias]
        mapa_ag = {a['nombre']: a['id_agencia'] for a in agencias}
        
        prov_final = col1.selectbox("🏢 Agencia / Partner Responsable", ["--- Seleccione ---"] + nombres_ag)
        
        is_disabled = bool(id_itinerario_dig)
        pax_name = col1.text_input("Pasajero Principal", value=def_pax, disabled=False)
        tel_pax = col1.text_input("Celular Contacto", value=def_cel)
        
        # Campo de Fecha de Venta (Editable, default hoy)
        fecha_venta_sel = col1.date_input("Fecha de Venta", value=date.today(), key="b2b_fecha_venta", help="La fecha en la que se realizó/concretó esta venta B2B.")
        
        vendedor_log = st.session_state.get('user_id', 'Operaciones')
        col1.markdown(f"👤 **Vendedor Resp:** {vendedor_log}")

        tour_name = col2.text_input("Nombre del Programa B2B", value=def_tour, disabled=False)
        tipo_comp = col2.radio("Comprobante para Agencia", ["Boleta", "Factura", "Recibo Simple"], horizontal=True, key="b2b_tipo_comp_v2")
        metodo_pago = col2.selectbox("💳 Método de Pago", ["EFECTIVO", "TRANSFERENCIA", "YAPE", "PLIN", "TARJETA", "PAYPAL", "OTRO"], key="b2b_metodo_pago_v2")
        
        cantidad_pax = col1.number_input("Cantidad Pax", min_value=1, value=int(def_cant_pax), disabled=is_disabled, key="b2b_pax_v2")

        # --- 📅 MOSTRAR FECHAS Y PAX (AL ESTILO B2C) ---
        if id_itinerario_dig:
            st.success(f"🗓️ **Viaje Programado:** Del {def_f_inicio.strftime('%d/%m/%Y')} al {def_f_fin.strftime('%d/%m/%Y')} | 👥 **Pax:** {def_cant_pax}")
            fecha_inicio_sel = def_f_inicio
            fecha_fin_sel = def_f_fin
        else:
            c_f1, c_f2 = st.columns(2)
            fecha_inicio_sel = c_f1.date_input("Fecha Inicio", value=def_f_inicio, key="b2b_finicio")
            fecha_fin_sel = c_f2.date_input("Fecha Fin", value=def_f_fin, key="b2b_ffin")

        # --- DESGLOSE DE INGRESOS B2B (MÁS ROBUSTO) ---
        items_ingreso = []
        if id_itinerario_dig:
            cached_items_b2b = st.session_state.get(f"b2b_items_{id_itinerario_dig}", [])
            if cached_items_b2b:
                for it in cached_items_b2b:
                    desc_b2b = it['descripcion']
                    if it['tipo'] in ['EXTRANJERO', 'CAN']:
                        desc_b2b += f" (Ref: ${it['p_raw']:.2f} x {tc_itin})"
                    
                    items_ingreso.append({
                        "descripcion": desc_b2b,
                        "cantidad": it['cantidad'],
                        "precio_unitario": it['precio_unitario']
                    })
                    st.info(f"✨ **{desc_b2b}**: Se han cargado **{it['cantidad']}** pax a **S/ {it['precio_unitario']:,.2f}** c/u.")
            else:
                st.warning("⚠️ No se pudo procesar el desglose B2B automático.")
        else:
            st.caption("No hay itinerario vinculado. El desglose se generará automáticamente por el total.")
        
        st.markdown("##### 🛡️ Logística y Emergencia (Pasajero Principal)")
        col_log1, col_log2 = st.columns(2)
        vuelo_int = col_log1.text_input("Nro de Vuelo Internacional", placeholder="Ej: AA947 / LA2345")
        correo_cli = col_log2.text_input("Correo Electrónico", placeholder="ejemplo@correo.com")
        
        cont_nom = col_log1.text_input("Nombre del Contacto de Emergencia", placeholder="Ej: María García (Hermana)")
        cont_tel = col_log2.text_input("Teléfono del Contacto de Emergencia", placeholder="+51 999 888 777")

        comentarios_op = st.text_area("🗒️ Comentarios para Operaciones", placeholder="Ej: Pasajero alérgico, requiere recojo puntual, etc.", key="coment_op_b2b")
        
        st.markdown("##### 📧 Notificaciones y Adjuntos B2B")
        c_not_b1, c_not_b2 = st.columns([1, 1])
        enviar_notif_b2b = c_not_b1.checkbox("Enviar Resumen por Correo Corporativo", value=True, help="Envía un resumen de la venta B2B a los correos de gerencia y reservas.")
        archivos_adjuntos_b2b = c_not_b2.file_uploader("Adjuntar Comprobantes o Fotos", accept_multiple_files=True, help="Opcional: Estas imágenes se enviarán como adjuntos en el correo.", key="uploader_b2b")

        st.divider()
        submitted = st.form_submit_button("✅ REGISTRAR VENTA B2B Y NOTIFICAR", use_container_width=True, type="primary")

        if submitted:
            if prov_final == "--- Seleccione ---":
                st.error("❌ Debe seleccionar una Agencia/Partner.")
            elif not pax_name or not tour_name:
                st.error("❌ El nombre del pasajero y del programa son obligatorios.")
            elif monto_total <= 0:
                st.error("❌ El monto total debe ser mayor a 0.")
            else:
                id_age = mapa_ag.get(prov_final)
                exito, msg = venta_controller.registrar_venta_proveedor(
                    nombre_proveedor=prov_final,
                    nombre_cliente=pax_name,
                    telefono=tel_pax,
                    vendedor=vendedor_log,
                    tour=tour_name,
                    monto_total=monto_total,
                    monto_depositado=monto_pagado,
                    id_agencia_aliada=id_age,
                    fecha_inicio=fecha_inicio_sel,
                    fecha_fin=fecha_fin_sel,
                    cantidad_pax=int(cantidad_pax),
                    id_itinerario_digital=id_itinerario_dig,
                    id_lead=None,
                    tipo_comprobante=tipo_comp,
                    tipo_cambio=tipo_cambio,
                    items_ingreso=items_ingreso if items_ingreso else None,
                    metodo_pago=metodo_pago,
                    vuelo_internacional=vuelo_int,
                    correo=correo_cli,
                    contacto_emergencia_nombre=cont_nom,
                    contacto_emergencia_tel=cont_tel,
                    comentarios=comentarios_op,
                    enviar_correo=enviar_notif_b2b,
                    adjuntos={f.name: f.getvalue() for f in archivos_adjuntos_b2b} if archivos_adjuntos_b2b else None,
                    fecha_venta=fecha_venta_sel.isoformat()
                )
                
                if exito:
                    st.success(f"🚀 {msg}")
                    st.balloons()
                else:
                    st.error(msg)

def reporte_operativo(controller):
    """Vista global de operaciones (Dashboard + Detalle)."""
    st.subheader("📊 Reporte Operativo Global", divider='blue')
    
    # 1. Dashboard de Analítica (Top)
    from vistas.dashboard_analytics import render_operations_dashboard
    df_ops = controller.get_data_for_analytics()
    render_operations_dashboard(df_ops)
    
    st.divider()
    
    # 2. Detalle de Operaciones (Auditoría)
    st.write("### 📋 Detalle de Servicios Programados")
    # Traer todos los servicios (no solo los de un día)
    range_start = date.today() - timedelta(days=30)
    range_end = date.today() + timedelta(days=60)
    todos_servicios = controller.get_servicios_rango_fechas(range_start, range_end)
    
    if not todos_servicios:
        st.info("No hay servicios registrados en el rango de tiempo seleccionado.")
    else:
        df_all = pd.DataFrame(todos_servicios)
        st.dataframe(
            df_all,
            column_order=("Fecha", "Servicio", "Tipo", "Pax", "Cliente", "Guía", "Estado Pago"),
            column_config={
                "Fecha": st.column_config.DateColumn("Fecha"),
                "Tipo": st.column_config.TextColumn("Tipo", width="small"),
                "Estado Pago": st.column_config.TextColumn("Pago")
            },
            hide_index=True,
            use_container_width=True
        )
        
        # 3. Vista Previa de Itinerario (Estilo Imagen)
        st.markdown("---")
        st.subheader("🏁 Verificador de Inclusiones (Itinerario)")
        
        if 'ID Itinerario' in df_all.columns:
            ventas_con_itin = df_all[df_all['ID Itinerario'].notna()]
            if not ventas_con_itin.empty:
                sel_id_v = st.selectbox("Auditar Itinerario de la Venta:", 
                                      ventas_con_itin['ID Venta'].unique(),
                                      format_func=lambda x: f"{ventas_con_itin[ventas_con_itin['ID Venta']==x]['Cliente'].values[0]} ({x})",
                                      key="sb_ops_audit_it")
                
                # Reutilizar lógica de detalle visual
                df_match = df_all[df_all['ID Venta'] == sel_id_v]
                id_itin_audit = df_match['ID Itinerario'].dropna().unique()[0] if not df_match['ID Itinerario'].dropna().empty else None
                
                if id_itin_audit:
                    res = controller.client.table('itinerario_digital').select('datos_render').eq('id_itinerario_digital', id_itin_audit).single().execute()
                    if res.data:
                        render_data = res.data['datos_render']
                        if isinstance(render_data, str):
                            import json
                            render_data = json.loads(render_data)
                            
                        if isinstance(render_data, dict):
                            # Enriquecer con datos de la fila de auditoría
                            dr = ventas_con_itin[ventas_con_itin['ID Venta'] == sel_id_v].iloc[0]
                            render_data['num_pasajeros'] = dr.get('num_pasajeros') or dr.get('Pax', 1)
                            render_data['ninos'] = dr.get('ninos', 0)
                            
                            # Rescate de nombre seguro
                            new_name = dr.get('Cliente')
                            if new_name and str(new_name).strip() not in ["", "---", "None", "Desconocido"]:
                                render_data['nombre_pasajero'] = new_name
                            
                            # Enriquecer con celular
                            render_data['cliente_telefono'] = dr.get('Celular') or ""
                                
                            render_itinerary_simple_download(render_data)
            else:
                st.info("Seleccione un servicio con itinerario para ver su detalle.")

def mostrar_pagina(nombre_modulo, rol_actual, user_id, supabase_client):
    """Punto de entrada de Streamlit para el área de Operaciones."""
    import controllers.operaciones_controller
    import controllers.venta_controller
    import models.venta_model
    import importlib
    
    # Forzar recarga de módulos para captar cambios en clases
    importlib.reload(models.venta_model)
    importlib.reload(controllers.operaciones_controller)
    importlib.reload(controllers.venta_controller)
    
    controller = controllers.operaciones_controller.OperacionesController(supabase_client)
    
    st.title("⚙️ Gestión de Operaciones")
    
    # --- 🔔 NUEVO: CENTRO DE ALERTAS ---
    render_centro_alertas(controller)
    
    st.markdown("---")
    
    if nombre_modulo in ["Gestión de Registros", "Logística y Proveedores", "Gestión de Operaciones"]:
        tab1, tab2, tab3 = st.tabs([
            "📊 Estructurador de Gastos (Master Sheet)",
            "🤝 Ventas B2B (Entrada)",
            "🏢 Directorio de Proveedores"
        ])
        
        with tab1:
            dashboard_simulador_costos(controller)
            
        with tab2:
            registro_ventas_proveedores(supabase_client)

        with tab3:
            render_directorio_proveedores(supabase_client)

    elif nombre_modulo == "Dashboard Diario":
        dashboard_tablero_diario(controller)
    elif nombre_modulo == "Reporte Operativo":
        reporte_operativo(controller)
    else:
        st.info("Seleccione una opción válida del menú lateral.")
            
# Función dashboard_pasajeros eliminada por simplificación de procesos
            
            


def dashboard_simulador_costos(controller):
    """
    Herramienta avanzada para estructurar liquidaciones de grupos/B2B.
    Basado en estructura de Excel (Unitario x Pax = Total).
    """
    st.subheader("📊 Panel de Control Profesional", divider='rainbow')

    # ✅ INICIALIZAR VARIABLES DE SESIÓN CRÍTICAS
    if 'last_loaded_id_venta' not in st.session_state:
        st.session_state['last_loaded_id_venta'] = None
    if 'master_pax_count' not in st.session_state:
        st.session_state['master_pax_count'] = 1

    # Pre-cargar proveedores para evitar errores de scope
    prov_items = []
    try:
        res_prov = controller.client.table('proveedor').select('id_proveedor, nombre_comercial, servicios_ofrecidos').execute()
        prov_items = res_prov.data or []
    except Exception as e:
        print(f"Error cargando proveedores init: {e}")
    
    # Inicializar variables de scope
    ventas_age = []
    mapa_ventas_pax = {}
    pax_sel = "--- Seleccione ---"
    all_edited = pd.DataFrame()

    if 'simulador_data' not in st.session_state:
        st.session_state['simulador_data'] = [
            {"FECHA": date.today(), "SERVICIO": "Servicio Ejemplo", "MONEDA": "USD", "TOTAL": 0.0},
        ]

    st.info("💡 Selecciona el tipo de venta y luego la venta específica para cargar sus datos.")
    
    # Barra de Agencias (Existente)
    from controllers.venta_controller import VentaController
    vc = VentaController(controller.client)
    agencias = vc.obtener_agencias_aliadas()
    nombres_agencias = [a['nombre'] for a in agencias]
    mapa_agencias = {a['nombre']: a['id_agencia'] for a in agencias}
    
    # PASO 1: Seleccionar Tipo de Venta
    c_tipo, c_filtro, c_pax = st.columns([1, 2, 2])
    
    with c_tipo:
        tipo_venta = st.selectbox("1️⃣ Tipo de Venta:", ["--- Seleccione ---", "🏢 B2B (Agencias)", "👤 B2C (Directas)"], key="sel_tipo_venta")
    
    ventas_age = []
    agencia_sel = None
    
    # PASO 2: Filtro según tipo
    if tipo_venta == "🏢 B2B (Agencias)":
        with c_filtro:
            agencia_sel = st.selectbox("2️⃣ Seleccione Agencia:", ["--- Seleccione ---"] + nombres_agencias, key="sel_agencia_b2b")
        
        if agencia_sel != "--- Seleccione ---":
            id_ag = mapa_agencias.get(agencia_sel)
            ventas_age = vc.obtener_ventas_agencia(id_ag)
    
    elif tipo_venta == "👤 B2C (Directas)":
        with c_filtro:
            st.info("📋 Mostrando todas las ventas directas")
        ventas_age = vc.obtener_ventas_directas()
    
    # PASO 3: Seleccionar Venta Específica
    if ventas_age:
        # Excluir ventas aprobadas por gerencia para operaciones
        ventas_age = [v for v in ventas_age if not v.get('aprobado_gerencia', False)]
        
        opciones_pax = [f"{v['nombre_cliente']} | {v.get('tour_nombre', 'Sin Tour')} ({v['id_venta']})" for v in ventas_age]
        mapa_ventas_pax = {f"{v['nombre_cliente']} | {v.get('tour_nombre', 'Sin Tour')} ({v['id_venta']})": v for v in ventas_age}
        
        with c_pax:
            pax_sel = st.selectbox("3️⃣ Cargar Venta:", ["--- Seleccione ---"] + opciones_pax, key="sel_pax_sim")
        
        if pax_sel != "--- Seleccione ---":
            v = mapa_ventas_pax.get(pax_sel)
            
            # Solo cargar si ha cambiado la venta para evitar bucles de rerun
            if st.session_state.get('last_loaded_id_venta') != v['id_venta']:
                st.session_state['master_pax_count'] = v.get('num_pasajeros', 1)
                
                # Actualizar venta actual
                st.session_state['last_loaded_id_venta'] = v['id_venta']
                st.rerun()

    # Carga de Archivos
    id_venta_act = st.session_state.get('last_loaded_id_venta')
    if not id_venta_act:
        st.info("Seleccione una venta para gestionar su información.")
        return

    # --- NUEVO: RECUPERAR DATOS DEL ITINERARIO PARA DESCARGA ---
    try:
        # Recuperar Venta Live de forma estándar (sin alias complejos para evitar fallos de PostgREST)
        res_v_live = controller.client.table('venta').select('*, cliente(nombre, lead(numero_celular, pais_origen))').eq('id_venta', id_venta_act).single().execute()
        v_live = res_v_live.data or {}
        
        id_itin_dig = v_live.get('id_itinerario_digital')
        render_data = None
        try:
            if id_itin_dig:
                res_render = controller.client.table('itinerario_digital').select('datos_render').eq('id_itinerario_digital', id_itin_dig).single().execute()
                if res_render.data:
                    render_data = res_render.data.get('datos_render')
                    if isinstance(render_data, str):
                        import json
                        render_data = json.loads(render_data)
                    
                    if isinstance(render_data, dict):
                        # Rescate de nombre infalible
                        cliente_obj = v_live.get('cliente') or {}
                        # PostgREST a veces devuelve el objeto directo o en una lista de 1 elemento
                        nombre_raw = "CLIENTE"
                        telefono_raw = ""
                        
                        if isinstance(cliente_obj, dict):
                            nombre_raw = cliente_obj.get('nombre')
                            lead_data = cliente_obj.get('lead')
                            if isinstance(lead_data, dict):
                                telefono_raw = lead_data.get('numero_celular', '')
                            elif isinstance(lead_data, list) and len(lead_data) > 0:
                                telefono_raw = lead_data[0].get('numero_celular', '')
                        elif isinstance(cliente_obj, list) and len(cliente_obj) > 0:
                            nombre_raw = cliente_obj[0].get('nombre')
                            lead_data = cliente_obj[0].get('lead')
                            if isinstance(lead_data, dict):
                                telefono_raw = lead_data.get('numero_celular', '')
                            elif isinstance(lead_data, list) and len(lead_data) > 0:
                                telefono_raw = lead_data[0].get('numero_celular', '')
                        
                        render_data['cliente_telefono'] = telefono_raw
                        
                        # Si sigue sin haber nombre, usar el del itinerario o un fallback genérico
                        nombre_final = nombre_raw or render_data.get('nombre_pasajero') or "CLIENTE"
                        
                        # Limpieza final: si es "---" o similar, forzar a "CLIENTE"
                        if str(nombre_final).strip() in ["", "---", "None", "Desconocido"]:
                            nombre_final = "CLIENTE"
                            
                        render_data['nombre_pasajero'] = str(nombre_final).upper()
                        render_data['num_pasajeros'] = v_live.get('num_pasajeros', 1)
                        render_data['ninos'] = v_live.get('ninos', 0)
                        render_data['fecha_inicio'] = v_live.get('fecha_inicio')
                        render_data['fecha_fin'] = v_live.get('fecha_fin')
                        
                        # Carga tours en vivo para sincronizar fechas
                        res_vt = controller.client.table('venta_tour').select('fecha_servicio').eq('id_venta', id_venta_act).order('n_linea').execute()
                        live_tours = res_vt.data or []
                        
                        itin_list = (render_data.get('itinerario_detalles') or 
                                     render_data.get('days') or 
                                     render_data.get('itinerario') or [])
                        
                        # Asegurar que itin_list es una lista mutable
                        if isinstance(itin_list, list):
                            for i, tour_live in enumerate(live_tours):
                                if i < len(itin_list) and isinstance(itin_list[i], dict):
                                    # Sincronizar ÚNICAMENTE la fecha operativa para el reporte
                                    itin_list[i]['fecha'] = tour_live.get('fecha_servicio') or itin_list[i].get('fecha')
                            
                            render_data['itinerario_detalles'] = itin_list
                    
                    # Renderizar los botones de descarga
                    if render_data:
                        with st.expander("📄 Ver Resumen de Itinerario (Simplificado)", expanded=False):
                            render_itinerary_simple_download(render_data)
        except Exception as e:
            st.warning(f"Nota: No se pudo cargar el resumen del itinerario PDF/Simple. ({e})")
        
        # --- SECCIÓN: EXPEDIENTE DIGITAL (DRIVE) ---
        st.markdown("### 📂 Expediente Digital")
        c_drive_1, c_drive_2 = st.columns([4, 1])
        with c_drive_1:
            url_actual = v_live.get('drive_url') or ""
            url_nueva = st.text_input("Link a Carpeta de Google Drive:", value=url_actual, placeholder="Pegue aquí el enlace del Drive del pasajero", key="drive_input")
        with c_drive_2:
            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("💾 Guardar Link", use_container_width=True):
                try:
                    controller.client.table('venta').update({'drive_url': url_nueva}).eq('id_venta', id_venta_act).execute()
                    st.toast("✅ Enlace de Drive actualizado!")
                    st.rerun()
                except Exception as e:
                    st.error(f"Error al guardar link: {e}")
        
        if url_actual:
            st.link_button("👉 Abrir Carpeta de Archivos (Drive)", url_actual, use_container_width=True, type="secondary")

        # --- NUEVO: EDITAR LOGÍSTICA DE VENTA ---
        st.markdown("### 🛡️ Información Logística y de Emergencia")
        with st.expander("📝 Ver / Editar Datos de Logística", expanded=False):
            with st.form(f"form_logistica_op_{id_venta_act}"):
                col_log1, col_log2 = st.columns(2)
                v_nro = col_log1.text_input("Nro de Vuelo Internacional", value=v_live.get('nro_vuelo_internacional') or "")
                v_correo = col_log2.text_input("Correo Electrónico", value=v_live.get('correo_cliente') or "")
                v_cont = col_log1.text_input("Contacto de Emergencia", value=v_live.get('nombre_contacto_emergencia') or "")
                v_tel = col_log2.text_input("Teléfono de Emergencia", value=v_live.get('telefono_contacto_emergencia') or "")
                
                if st.form_submit_button("💾 Guardar Datos Logísticos", use_container_width=True):
                    try:
                        controller.client.table('venta').update({
                            'nro_vuelo_internacional': v_nro,
                            'correo_cliente': v_correo,
                            'nombre_contacto_emergencia': v_cont,
                            'telefono_contacto_emergencia': v_tel
                        }).eq('id_venta', id_venta_act).execute()
                        st.success("✅ Datos logísticos actualizados correctamente.")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Error al guardar logística: {e}")

        # --- NUEVO: IMPRIMIR VOUCHER DE RESERVA ---
        st.markdown("### 🎫 Documento de Viaje")
        with st.expander("🎫 Generar / Imprimir Voucher de Reserva", expanded=False):
            from vistas.componentes_voucher import render_panel_voucher
            render_panel_voucher(controller.client, id_venta_default=id_venta_act, key_prefix="ops_vch")

        # Botón Maestro Operativo (Independiente del Itinerario Digital)
        # Se pone fuera del bloque anterior para que funcione aunque no haya Itinerario JSON
        st.markdown("---")
        render_operational_master_download(controller, id_venta_act)
    except Exception as e:
        st.error(f"❌ Error crítico en sección de descargas: {e}")

    # --- SECCIÓN DE ARCHIVOS (CSV/Excel) ---
    st.markdown("### 📝 Gestión de Información Externa")
    st.info("Suba los archivos correspondientes para el cierre y control de pasajeros.")

    # NUEVO: Botón de Plantilla con Listas Desplegables
    import io
    import openpyxl
    from openpyxl.worksheet.datavalidation import DataValidation
    
    buffer = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plantilla"
    
    # Headers
    headers = ["Dia", "Hora", "Tipo de Servicio", "Proveedor", "Nombre del Guia", "Observacion", "Pax"]
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)
        
    # Obtener proveedores de la BD
    try:
        res_prov = controller.client.table('proveedor').select('nombre_comercial').execute()
        lista_provs = sorted(list(set([str(p['nombre_comercial']).strip() for p in (res_prov.data or []) if p.get('nombre_comercial')])))
    except:
        lista_provs = []
        
    if lista_provs:
        ws_listas = wb.create_sheet("Listas")
        ws_listas.sheet_state = 'hidden'
        for i, prov in enumerate(lista_provs, 1):
            ws_listas.cell(row=i, column=1, value=prov)
            
        # Validación de datos para Proveedor (Columna D = 4)
        dv_prov = DataValidation(type="list", formula1=f"Listas!$A$1:$A${len(lista_provs)}", allow_blank=True)
        ws.add_data_validation(dv_prov)
        dv_prov.add("D2:D100")
        
    wb.save(buffer)
    buffer.seek(0)
    
    st.download_button(
        label="📥 Descargar Plantilla Inteligente (Excel)",
        data=buffer.getvalue(),
        file_name="plantilla_endoses_inteligente.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    
    c_arch1, c_arch2 = st.columns(2)
    
    # ✅ VALIDACIÓN ANTES DE MOSTRAR UPLOADER
    id_venta_para_archivos = st.session_state.get('last_loaded_id_venta')
    if not id_venta_para_archivos:
        with c_arch1:
            st.warning("⚠️ Selecciona una venta arriba para cargar archivos.", icon="🔴")
        with c_arch2:
            st.warning("⚠️ Selecciona una venta arriba para cargar pasajeros.", icon="🔴")
        return
    
    with c_arch1:
        st.subheader("📊 Panel de Control")
        f_liq = st.file_uploader("Cierre de Operaciones (Excel/CSV):", type=['xlsx', 'xls', 'csv'], key="up_liqrar_final", help="Carga el archivo de endoses con columnas: Día, Tipo de Servicio, Proveedor")
        
        # FUNCIÓN AUXILIAR: Normalizar nombres de columnas
        def normalizar_columnas(df):
            """Convierte columnas a lowercase, sin espacios extras, para flexible matching"""
            df.columns = [str(col).strip().lower() for col in df.columns]
            return df
        
        def mapear_columnas_flexible(df_norm):
            """Mapea columnas normalizadas a las esperadas, siendo flexible"""
            mapeo = {
                'dia': ['dia', 'día', 'n_linea', 'nlinea', 'línea', 'linea', 'day'],
                'tipo_de_servicio': ['tipo_de_servicio', 'tipo de servicio', 'tipo_servicio', 'servicio', 'tipo', 'service_type'],
                'proveedor': ['proveedor', 'provider', 'supplier', 'nombre_proveedor', 'empresa']
            }
            
            cols_encontradas = {}
            for col_esperada, aliases in mapeo.items():
                for col_actual in df_norm.columns:
                    if col_actual in aliases or any(alias in col_actual for alias in aliases):
                        cols_encontradas[col_esperada] = col_actual
                        break
            
            return cols_encontradas
        
        # PROCESAMIENTO DE ENDOSES
        if f_liq:
            try:
                # Cargar datos
                if f_liq.name.endswith('.csv'):
                    df_preview = pd.read_csv(f_liq)
                else:
                    df_preview = pd.read_excel(f_liq)
                
                # Limpiar filas vacías
                df_preview = df_preview.dropna(how='all')
                
                if df_preview.empty:
                    st.error("❌ El archivo está vacío. Completa la plantilla y vuelve a intentar.")
                else:
                    # Previsualización
                    st.write("**Previsualización de datos a cargar:**")
                    st.dataframe(df_preview, use_container_width=True, hide_index=True)
                    
                    # Validar columnas con flexibilidad
                    cols_req = ["Dia", "Tipo de Servicio", "Proveedor"]
                    df_norm = normalizar_columnas(df_preview.copy())
                    cols_mapeadas = mapear_columnas_flexible(df_norm)
                    
                    if len(cols_mapeadas) == len(cols_req):
                        st.success(f"✅ Columnas detectadas correctamente: {list(cols_mapeadas.keys())}")
                        
                        # Renombrar columnas en el dataframe original para compatibilidad
                        rename_dict = {}
                        for col_esperada, col_actual in cols_mapeadas.items():
                            for idx, col in enumerate(df_preview.columns):
                                if str(col).strip().lower() == col_actual:
                                    # BUGFIX: Usar los nombres exactos esperados por vincular_endoses_masivos
                                    # 'dia' -> 'Dia', 'tipo_de_servicio' -> 'Tipo de Servicio', etc
                                    if col_esperada == 'dia':
                                        rename_dict[col] = 'Dia'
                                    elif col_esperada == 'tipo_de_servicio':
                                        rename_dict[col] = 'Tipo de Servicio'
                                    elif col_esperada == 'proveedor':
                                        rename_dict[col] = 'Proveedor'
                        
                        if rename_dict:
                            df_preview = df_preview.rename(columns=rename_dict)
                            print(f"DEBUG - Columnas renombradas: {rename_dict}")  # ✅ LOGGING
                        
                        # --- CONFIGURACIÓN DE TIPO DE CAMBIO PARA LA CARGA ---
                        tc_carga = st.number_input("💱 Tipo de Cambio para esta Carga (USD -> PEN):", min_value=0.1, value=3.80, format="%.3f", help="Se usará para convertir costos si la moneda del proveedor es distinta a la de la venta.")
                        
                        if st.button("📦 Procesar y Guardar Endoses en DB", type="primary", use_container_width=True):
                            # ✅ VALIDACIÓN FINAL ANTES DE PROCESAR
                            if not id_venta_para_archivos:
                                st.error("❌ Error interno: No hay venta seleccionada. Recarga la página e intenta de nuevo.")
                                st.stop()
                            
                            try:
                                print(f"\n{'='*60}")
                                print(f"🔵 INICIANDO CARGA DE ENDOSES")
                                print(f"ID VENTA: {id_venta_para_archivos}")
                                print(f"FILAS A PROCESAR: {len(df_preview)}")
                                print(f"COLUMNAS: {list(df_preview.columns)}")
                                print(f"TIPO DE CAMBIO: {tc_carga}")
                                print(f"{'='*60}\n")
                                
                                with st.spinner("⏳ Procesando archivo... esto puede tomar unos segundos"):
                                    print(f"🟡 Llamando a vincular_endoses_masivos...")
                                    res_bulk = controller.vincular_endoses_masivos(id_venta_para_archivos, df_preview, tc_manual=tc_carga)
                                    print(f"🔵 Respuesta recibida: {res_bulk}")
                                    print(f"Tipo de respuesta: {type(res_bulk)}")
                                
                                # ✅ VALIDACIÓN MEJORADA DE RESPUESTA
                                if not isinstance(res_bulk, dict):
                                    st.error(f"❌ ERROR CRÍTICO: Respuesta inválida del servidor. Tipo: {type(res_bulk)}")
                                    print(f"❌ ERROR: res_bulk no es diccionario: {res_bulk}")
                                    st.stop()
                                
                                exitos = res_bulk.get('exitos', 0)
                                errores = res_bulk.get('errores', [])
                                
                                print(f"\n📊 RESULTADO:")
                                print(f"  Éxitos: {exitos}")
                                print(f"  Errores: {len(errores)}")
                                for e in errores[:5]:  # Mostrar primeros 5 errores
                                    print(f"    - {e}")
                                
                                if exitos == 0 and not errores:
                                    st.error("❌ La operación completó sin guardar datos. Verifica que la venta tenga itinerario sincronizado.")
                                    st.info("💡 Datos de diagnóstico (para tu equipo técnico):\n" + 
                                           f"- ID Venta: {id_venta_para_archivos}\n" +
                                           f"- Filas en archivo: {len(df_preview)}\n" +
                                           f"- Respuesta: {res_bulk}")
                                elif exitos > 0:
                                    st.success(f"✅ Se vincularon {exitos} registros correctamente.")
                                    if errores:
                                        st.warning(f"⚠️ Se encontraron {len(errores)} advertencias:")
                                        with st.expander("📋 Ver detalles de errores", expanded=False):
                                            for err in errores:
                                                st.write(f"• {err}")
                                    st.balloons()
                                    print(f"\n✅ ÉXITO - Guardados {exitos} registros\n")
                                    st.rerun()
                                else:
                                    if errores:
                                        st.error(f"❌ No se guardó ningún registro. Errores encontrados ({len(errores)}):")
                                        for err in errores[:10]:  # Mostrar hasta 10 errores
                                            st.write(f"• {err}")
                                    else:
                                        st.error("❌ Error desconocido al procesar el archivo.")
                            except Exception as process_error:
                                print(f"\n{'='*60}")
                                print(f"❌ EXCEPCIÓN EN PROCESAMIENTO:")
                                print(f"{type(process_error).__name__}: {str(process_error)}")
                                import traceback
                                print(traceback.format_exc())
                                print(f"{'='*60}\n")
                                
                                st.error(f"❌ Error al procesar: {str(process_error)}")
                                st.info("💡 Detalles técnicos (para tu equipo):\n```\n" + 
                                       f"{type(process_error).__name__}: {str(process_error)}\n```")

                                st.info("💡 Verifica que:")
                                st.write("• La venta tiene un itinerario sincronizado")
                                st.write("• Los proveedores del archivo existen en el sistema")
                                st.write("• Los días (1,2,3...) corresponden a los días del itinerario")
                    else:
                        # Mostrar diagnóstico detallado
                        st.error(f"❌ No se encontraron todas las columnas requeridas")
                        st.info("**Columnas esperadas:**")
                        st.write(f"• `Dia` (o 'Día', 'N Linea', 'Línea', 'Day')")
                        st.write(f"• `Tipo de Servicio` (o 'Tipo Servicio', 'Servicio', 'Service')")
                        st.write(f"• `Proveedor` (o 'Provider', 'Supplier', 'Empresa')")
                        
                        st.info("**Columnas detectadas en tu archivo:**")
                        st.write(f"{', '.join(df_preview.columns.tolist())}")
                        
                        st.info("💡 Descarga la plantilla predefinida y complétala para evitar errores de formato.")
            except Exception as e:
                st.error(f"❌ Error al leer el archivo: {str(e)}")
                st.info("💡 Asegúrate de que:")
                st.write("• El archivo no está corrupto")
                st.write("• Es un Excel (.xlsx, .xls) o CSV válido")
                st.write("• Contiene al menos una fila de datos")
        else:
            st.info("📁 Carga un archivo de endoses para procesarlo")

    with c_arch2:
        st.subheader("👥 Pasajeros")
        # Botón de Plantilla Pax
        pax_template_df = pd.DataFrame(columns=['Nombre', 'Apellidos', 'Documento', 'Tipo Doc', 'Fecha Caducidad', 'Nacionalidad', 'Fecha Nacimiento', 'Edad', 'Genero', 'Tipo Habitacion', 'Cuidados', 'Es Principal'])
        pax_buffer = io.BytesIO()
        with pd.ExcelWriter(pax_buffer, engine='openpyxl') as writer:
            pax_template_df.to_excel(writer, index=False, sheet_name='Rooming')
        
        st.download_button(
            label="📥 Descargar Plantilla Rooming",
            data=pax_buffer.getvalue(),
            file_name="plantilla_rooming.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

        f_pax = st.file_uploader("Lista de Pasajeros / Rooming (Excel/CSV):", type=['xlsx', 'xls', 'csv'], key="up_paxrar_final", help="Carga el rooming con información de los pasajeros")
        if f_pax:
            try:
                if f_pax.name.endswith('.csv'):
                    df_pax = pd.read_csv(f_pax)
                else:
                    df_pax = pd.read_excel(f_pax)
                
                st.dataframe(df_pax, use_container_width=True, hide_index=True)
                
                if st.button("👥 Cargar Rooming a la DB", type="primary", use_container_width=True):
                    # ✅ VALIDACIÓN FINAL ANTES DE PROCESAR
                    if not id_venta_para_archivos:
                        st.error("❌ Error interno: No hay venta seleccionada. Recarga la página e intenta de nuevo.")
                        st.stop()
                    
                    try:
                        with st.spinner("⏳ Cargando pasajeros..."):
                            res_pax = controller.vincular_pasajeros_masivos(id_venta_para_archivos, df_pax)
                        
                        exitos_pax = res_pax.get('exitos', 0)
                        errores_pax = res_pax.get('errores', [])
                        
                        if exitos_pax > 0:
                            st.success(f"✅ Se cargaron {exitos_pax} pasajeros correctamente.")
                            if errores_pax:
                                with st.expander("⚠️ Advertencias", expanded=False):
                                    for e in errores_pax:
                                        st.warning(e)
                            st.rerun()
                        else:
                            st.error("❌ No se guardó ningún pasajero.")
                            if errores_pax:
                                with st.expander("📋 Ver errores"):
                                    for e in errores_pax:
                                        st.write(f"• {e}")
                    except Exception as pax_error:
                        st.error(f"❌ Error al cargar pasajeros: {str(pax_error)}")
            except Exception as e:
                st.error(f"❌ Error al leer archivo de pasajeros: {e}")

    st.divider()

    # --- SECCIÓN: RESUMEN DE ASIGNACIONES (VISUALIZACIÓN) ---
    c_res1, c_res2 = st.columns(2)
    
    with c_res1:
        st.markdown("### 📋 Panel de Control")
        try:
            id_actual = st.session_state.get('last_loaded_id_venta')
            # 1. Obtener Itinerario base (Días) y Moneda/TC de la Venta
            res_v = controller.client.table('venta').select('moneda, tipo_cambio').eq('id_venta', id_actual).single().execute()
            v_meta = res_v.data or {"moneda": "USD", "tipo_cambio": 3.8}
            moneda_v = v_meta.get('moneda', 'USD')
            tc_v = float(v_meta.get('tipo_cambio') or 3.8)

            servicios_v = vc.obtener_detalles_itinerario_venta(id_actual)
            mapa_servicios = {s['n_linea']: s for s in servicios_v}
            
            # 2. Obtener Liquidaciones Reales (Lo que se subió por Excel)
            liq_data = controller.get_liquidaciones_venta(id_actual)
            
            if liq_data:
                # Obtener lista de proveedores activos para el dropdown
                res_prov_drop = controller.client.table('proveedor').select('id_proveedor, nombre_comercial').eq('activo', True).order('nombre_comercial').execute()
                lista_proveedores = []
                mapa_proveedores_id = {}
                if res_prov_drop.data:
                    lista_proveedores = [p['nombre_comercial'] for p in res_prov_drop.data]
                    mapa_proveedores_id = {p['nombre_comercial']: p['id_proveedor'] for p in res_prov_drop.data}

                # 1. Preparar datos para el editor
                display_data = []
                for l in liq_data:
                    n_lin = l.get('n_linea')
                    
                    l['Dia'] = n_lin
                    l['Hora'] = l.get('hora_servicio') or "08:00 AM"
                    l['Tipo de Servicio'] = l.get('tipo_servicio', '---')
                    l['Proveedor'] = l.get('proveedor', {}).get('nombre_comercial') if l.get('proveedor') else "---"
                    l['Guía'] = l.get('nombre_guia', '---')
                    f_raw = l.get('fecha_confirmacion')
                    try:
                        l['F. Confirmación'] = pd.to_datetime(f_raw).date() if f_raw else None
                    except:
                        l['F. Confirmación'] = None
                    l['Resp. Contrato'] = l.get('responsable_contratacion', '')
                    l['Observacion'] = l.get('observacion', '---')
                    
                    # FECHAS Y CHECKS NUEVOS
                    f_cont = l.get('fecha_contratacion')
                    try:
                        l['F. Contratación'] = pd.to_datetime(f_cont).date() if f_cont else None
                    except:
                        l['F. Contratación'] = None
                    
                    l['Estado Contrato'] = "🟢 PAGADO" if l.get('contratado') else "⚪ PENDIENTE"
                    
                    # Mantener cálculos internos por si se usan luego (pueden estar ocultos)
                    c_unit = float(l.get('costo_unitario', 0))
                    pax = float(l.get('cantidad_pax') or l.get('cantidad_items') or 1)
                    l['PAX'] = int(pax)
                    l['TC'] = l.get('tipo_cambio') or tc_v
                    l['TOTAL (PEN)'] = (c_unit * pax) * float(l['TC']) if l.get('moneda') == 'USD' else (c_unit * pax)
                    
                    # ICONO DE ESTADO
                    l['Estado'] = "🟡 OK" if l.get('terminado') else "🔴 PENDIENTE"
                    display_data.append(l)

                df_edit = pd.DataFrame(display_data)
                
                # Definir columnas visibles: Logística + Finanzas
                cols_visible = [
                    'Estado', 'terminado', 
                    'Estado Contrato', 'contratado', 'F. Contratación',
                    'Dia', 'Hora', 'Tipo de Servicio', 'Proveedor', 'Guía', 'F. Confirmación', 'Resp. Contrato',
                    'metodo_pago', 'observaciones_pago', 'observaciones_contables',
                    'Observacion', 'moneda', 'costo_unitario', 'PAX', 'TC', 'TOTAL (PEN)'
                ]
                
                # 2. Renderizar Editor de Datos
                editor_key = f"editor_liq_master_{id_actual}"
                
                edited_result = st.data_editor(
                    df_edit[cols_visible],
                    column_config={
                        "Estado": st.column_config.TextColumn("Visual", width="small"),
                        "terminado": st.column_config.CheckboxColumn("Check", help="Marcar como Confirmado"),
                        "Estado Contrato": st.column_config.TextColumn("Visual Pago", width="small"),
                        "contratado": st.column_config.CheckboxColumn("Check Pago", help="Marcar como Pagado"),
                        "F. Contratación": st.column_config.DateColumn("F. Pago", width="small"),
                        "Dia": st.column_config.NumberColumn("Día", format="%d", width="small"),
                        "Hora": st.column_config.TextColumn("Hora", width="small"),
                        "Tipo de Servicio": st.column_config.TextColumn("Tipo de Servicio", width="medium"),
                        "Proveedor": st.column_config.SelectboxColumn("Proveedor", options=lista_proveedores, width="medium"),
                        "Guía": st.column_config.TextColumn("Guía", width="medium"),
                        "TC": st.column_config.NumberColumn("TC", format="%.3f", width="small"),
                        "F. Confirmación": st.column_config.DateColumn("Confirmación", width="small"),
                        "Resp. Contrato": st.column_config.TextColumn("Resp. Contrato", width="medium"),
                        "metodo_pago": st.column_config.SelectboxColumn("Método Pago", options=["---", "YAPE", "PLIN", "TRANSFERENCIA", "EFECTIVO", "OTRO"], width="medium"),
                        "observaciones_pago": st.column_config.TextColumn("Observación Pago", width="medium"),
                        "observaciones_contables": st.column_config.TextColumn("Obs. Contables", width="medium"),
                        "Observacion": st.column_config.TextColumn("Observación", width="large"),
                        "moneda": st.column_config.SelectboxColumn("Moneda", options=["USD", "PEN", "EUR"], width="small"),
                        "costo_unitario": st.column_config.NumberColumn("Costo Unit.", format="%.2f"),
                        "PAX": st.column_config.NumberColumn("Pax", width="small"),
                        "TOTAL (PEN)": st.column_config.NumberColumn("Costo Total (S/.)", format="S/. %.2f")
                    },
                    disabled=['Estado', 'Estado Contrato', 'TOTAL (PEN)'],
                    hide_index=True,
                    use_container_width=True,
                    num_rows="dynamic",
                    key=editor_key
                )
                guardar_btn = st.button("💾 Guardar Cambios en Operativa", type="primary", use_container_width=True, key=f"btn_guardar_op_{id_actual}")

                # 3. Procesar cambios mediante botón de confirmación
                if guardar_btn:
                    state = st.session_state.get(editor_key, {})
                    cambios_pendientes = state.get("edited_rows", {})
                    agregados_pendientes = state.get("added_rows", [])
                    borrados_pendientes = state.get("deleted_rows", [])
                    
                    if cambios_pendientes or agregados_pendientes or borrados_pendientes:
                        exitos = 0
                        errores = []
                        # Renombrar campos internos a nombres de DB
                        mapping = {
                            "moneda": "moneda", 
                            "costo_unitario": "costo_unitario", 
                            "PAX": "cantidad_pax", 
                            "TC": "tipo_cambio",
                            "terminado": "terminado",
                            "Hora": "hora_servicio",
                            "Guía": "nombre_guia",
                            "Observacion": "observacion",
                            "F. Confirmación": "fecha_confirmacion",
                            "F. Contratación": "fecha_contratacion",
                            "contratado": "contratado",
                            "Resp. Contrato": "responsable_contratacion",
                            "metodo_pago": "metodo_pago",
                            "observaciones_pago": "observaciones_pago",
                            "observaciones_contables": "observaciones_contables"
                        }
                        # Para nuevas filas solo se insertan columnas válidas en venta_servicio_proveedor.
                        insert_mapping = {k: v for k, v in mapping.items() if k not in ("metodo_pago", "observaciones_pago", "observaciones_contables")}
                        for row_idx, changes in cambios_pendientes.items():
                            reg_id = df_edit.iloc[row_idx]['id']
                            # Renombrar campos internos a nombres de DB y convertir fechas a string
                            db_changes = {}
                            for k, v in changes.items():
                                if k in mapping:
                                    val = v
                                    if hasattr(v, 'isoformat'): # Para objetos date de Streamlit
                                        val = v.isoformat()
                                    db_changes[mapping[k]] = val
                            
                            # --- LÓGICA DE VÍNCULO (Linkage) ---
                            # 1. Confirmación (terminado <-> fecha_confirmacion)
                            if 'terminado' in changes:
                                if changes['terminado'] is True:
                                    # Si se marca como OK y no hay fecha (ni en cambios ni en DF), poner hoy
                                    if not (changes.get('F. Confirmación') or df_edit.iloc[row_idx].get('F. Confirmación')):
                                        db_changes['fecha_confirmacion'] = date.today().isoformat()
                                else:
                                    # Si se desmarca, limpiar fecha
                                    db_changes['fecha_confirmacion'] = None
                            
                            if 'F. Confirmación' in changes:
                                if changes['F. Confirmación']:
                                    # Si se pone una fecha, marcar como OK
                                    db_changes['terminado'] = True
                                else:
                                    # Si se limpia la fecha, desmarcar OK
                                    db_changes['terminado'] = False

                            # 2. Contratación (contratado <-> fecha_contratacion)
                            if 'contratado' in changes:
                                if changes['contratado'] is True:
                                    # Si se marca como contratado y no hay fecha, poner hoy
                                    if not (changes.get('F. Contratación') or df_edit.iloc[row_idx].get('F. Contratación')):
                                        db_changes['fecha_contratacion'] = date.today().isoformat()
                                else:
                                    # Si se desmarca, limpiar fecha
                                    db_changes['fecha_contratacion'] = None
                            
                            if 'F. Contratación' in changes:
                                if changes['F. Contratación']:
                                    # Si se pone una fecha, marcar como contratado
                                    db_changes['contratado'] = True
                                else:
                                    # Si se limpia la fecha, desmarcar contratado
                                    db_changes['contratado'] = False
                            
                            # Lógica especial para Proveedor (Búsqueda por nombre)
                            if "Proveedor" in changes:
                                nom_prov = changes["Proveedor"]
                                if nom_prov in mapa_proveedores_id:
                                    db_changes['id_proveedor'] = mapa_proveedores_id[nom_prov]
                                else:
                                    # Buscar ID del proveedor por nombre comercial
                                    res_p = controller.client.table('proveedor').select('id_proveedor').ilike('nombre_comercial', f"%{nom_prov}%").limit(1).execute()
                                    if res_p.data:
                                        db_changes['id_proveedor'] = res_p.data[0]['id_proveedor']
                                    else:
                                        errores.append(f"⚠️ No se encontró el proveedor '{nom_prov}'. Se guardarán los demás cambios.")
                            
                            # La lógica de vinculación anterior ya cubre los casos de automatización
                             
                            if db_changes:
                                exito, msg = controller.actualizar_campos_liquidacion(reg_id, db_changes)
                                if exito: exitos += 1
                                else: errores.append(f"Error en fila {row_idx+1}: {msg}")
                        # --- PROCESAR FILAS AÑADIDAS (NUEVOS SERVICIOS) ---
                        exitos_nuevos = 0
                        for new_row in agregados_pendientes:
                            db_insert = {"id_venta": id_actual}
                            pago_fields = {}
                            
                            # Calcular N_linea (Día)
                            n_lin = new_row.get("Dia")
                            if not n_lin:
                                n_lin = df_edit['Dia'].max() + 1 if not df_edit.empty else 1
                            db_insert["n_linea"] = int(n_lin)
                            
                            # Aplicar mapeo solo a las columnas que pertenecen a venta_servicio_proveedor
                            for k, v in new_row.items():
                                if k in insert_mapping:
                                    val = v
                                    if hasattr(v, 'isoformat'): val = v.isoformat()
                                    db_insert[insert_mapping[k]] = val
                                elif k == "metodo_pago" and v:
                                    pago_fields["metodo_pago"] = v
                                elif k == "observaciones_pago":
                                    pago_fields["observaciones"] = v or ""
                                elif k == "observaciones_contables":
                                    pago_fields["observaciones_contables"] = v or ""
                                    
                            # Proveedor
                            if "Proveedor" in new_row and new_row["Proveedor"]:
                                nom_prov = new_row["Proveedor"]
                                if nom_prov in mapa_proveedores_id:
                                    db_insert['id_proveedor'] = mapa_proveedores_id[nom_prov]
                                else:
                                    res_p = controller.client.table('proveedor').select('id_proveedor').ilike('nombre_comercial', f"%{nom_prov}%").limit(1).execute()
                                    if res_p.data:
                                        db_insert['id_proveedor'] = res_p.data[0]['id_proveedor']
                                    else:
                                        errores.append(f"⚠️ Proveedor '{nom_prov}' no encontrado para nueva fila. Se creará sin proveedor asignado.")
                            
                            # Validaciones de BD
                            if 'costo_unitario' not in db_insert or db_insert['costo_unitario'] is None:
                                db_insert['costo_unitario'] = 0.0
                            
                            try:
                                # Asegurar que el día (n_linea) existe en el itinerario (venta_tour) para la Foreign Key
                                try:
                                    res_vt = controller.client.table('venta_tour').select('n_linea').eq('id_venta', id_actual).eq('n_linea', db_insert["n_linea"]).execute()
                                    if not res_vt.data:
                                        # Insertar fila de anclaje para este día
                                        controller.client.table('venta_tour').insert({
                                            'id_venta': id_actual,
                                            'n_linea': db_insert["n_linea"],
                                            'fecha_servicio': date.today().isoformat(),
                                            'observacion': 'Fila autogenerada para vinculación de costos adicionales'
                                        }).execute()
                                except Exception as e_vt:
                                    pass
                                    
                                controller.client.table('venta_servicio_proveedor').insert(db_insert).execute()
                                if pago_fields:
                                    monto_total = float(db_insert.get('costo_unitario', 0) or 0) * float(db_insert.get('cantidad_pax', 1) or 1)
                                    pago_insert = {
                                        "id_venta": id_actual,
                                        "n_linea": db_insert["n_linea"],
                                        "id_proveedor": db_insert.get('id_proveedor'),
                                        "monto_pagado": monto_total if monto_total > 0 else 0.0,
                                        "moneda": db_insert.get('moneda', 'USD') or 'USD',
                                        "monto_en_moneda_costo": monto_total if monto_total > 0 else 0.0,
                                        "fecha_pago": date.today().isoformat(),
                                        "metodo_pago": pago_fields.get('metodo_pago', 'TRANSFERENCIA'),
                                        "observaciones": pago_fields.get('observaciones', ''),
                                        "observaciones_contables": pago_fields.get('observaciones_contables', '')
                                    }
                                    controller.client.table('pago_operativo').insert(pago_insert).execute()
                                exitos_nuevos += 1
                            except Exception as e:
                                errores.append(f"Error al añadir fila (Día {n_lin}): {str(e)}")
                        # --- PROCESAR FILAS ELIMINADAS ---
                        exitos_borrados = 0
                        for row_idx in borrados_pendientes:
                            reg_id = df_edit.iloc[row_idx]['id']
                            try:
                                controller.client.table('venta_servicio_proveedor').delete().eq('id', reg_id).execute()
                                exitos_borrados += 1
                            except Exception as e:
                                errores.append(f"Error al eliminar fila {row_idx+1}: {str(e)}")
                        
                        if exitos > 0 or exitos_nuevos > 0 or exitos_borrados > 0:
                            st.success(f"✅ Se actualizaron {exitos}, se añadieron {exitos_nuevos} y se eliminaron {exitos_borrados} servicios.")
                            # Limpiar el editor forzando un reset
                            st.rerun()
                        if errores:
                            for e in errores: st.error(e)
                    confirm_reset = st.checkbox("Confirmar borrado de todos los costos (Liquidación)", key="reset_end_confirm")
                    if st.button("🗑️ Resetear Liquidación", type="primary", disabled=not confirm_reset, use_container_width=True):
                        exito_r, msg_r = controller.borrar_endoses_venta(id_actual)
                        if exito_r: st.success(msg_r); st.rerun()
                        else: st.error(msg_r)
            else:
                # Diagnóstico: diferenciar "no hay registros" vs "hay registros pero no están listos"
                try:
                    res_cnt = controller.client.table('venta_servicio_proveedor') \
                        .select('id', count='exact') \
                        .eq('id_venta', id_actual) \
                        .execute()
                    total_regs = getattr(res_cnt, "count", None)
                except Exception:
                    total_regs = None
                
                if total_regs and int(total_regs) > 0:
                    st.warning(
                        f"⚠️ Se detectaron {int(total_regs)} registros en la BD para esta venta, "
                        "pero no se pueden renderizar aún en el panel."
                    )
                    st.info("💡 Suele ocurrir si faltan campos clave (por ejemplo proveedor) o hubo un problema en la carga. Reintenta sincronizar y recarga.")
                else:
                    st.info("Aún no has cargado la liquidación (Excel) para esta venta.")
        except Exception as e:
            st.error(f"Error cargando liquidaciones: {e}")

    with c_res2:
        st.markdown("### 👥 Resumen de Pasajeros (Rooming)")
        try:
            pax_data = controller.pasajero_model.get_by_venta_id(id_actual)
            if pax_data:
                df_pax_res = pd.DataFrame(pax_data)
                
                # Definir columnas visibles y editables para logística
                # Mapeamos nombres internos a nombres amigables para el usuario
                cols_logistica = {
                    'nombre_completo': 'Pasajero',
                    'nacionalidad': 'Nac.',
                    'numero_documento': 'Doc',
                    'vuelo_llegada': '✈️ Llegada',
                    'vuelo_salida': '✈️ Salida',
                    'dieta': '🍽️ Dieta',
                    'acomodacion': '🛏️ Habitación',
                    'telefono': '📞 Teléfono',
                    'es_principal': 'P'
                }
                
                # Filtrar solo columnas que existan en el DF
                df_display = df_pax_res[[c for c in cols_logistica.keys() if c in df_pax_res.columns]]
                df_display.rename(columns=cols_logistica, inplace=True)

                st.info("💡 Puedes editar los vuelos, dietas y habitaciones directamente en la tabla y luego presionar 'Guardar Logística'.")
                
                edited_pax = st.data_editor(
                    df_display,
                    use_container_width=True,
                    hide_index=True,
                    key=f"editor_pax_log_{id_actual}",
                    column_config={
                        "P": st.column_config.CheckboxColumn("P", width="small")
                    }
                )

                # Lógica de guardado para el editor de pasajeros
                if st.session_state.get(f"editor_pax_log_{id_actual}"):
                    changes = st.session_state[f"editor_pax_log_{id_actual}"].get("edited_rows", {})
                    if changes:
                        st.warning(f"⚠️ Tienes {len(changes)} cambios de pasajeros pendientes.")
                        if st.button("💾 Guardar Cambios Logísticos", type="primary", use_container_width=True):
                            exitos = 0
                            for idx, row_changes in changes.items():
                                # Obtener el ID real del pasajero
                                real_pax_id = df_pax_res.iloc[idx]['id_pasajero']
                                
                                # Mapear nombres de vuelta
                                reverse_map = {v: k for k, v in cols_logistica.items()}
                                db_changes = {reverse_map[col]: val for col, val in row_changes.items()}
                                
                                try:
                                    controller.client.table('pasajero').update(db_changes).eq('id_pasajero', real_pax_id).execute()
                                    exitos += 1
                                except Exception as e:
                                    st.error(f"Error guardando pasajero {idx}: {e}")
                            
                            if exitos > 0:
                                st.success(f"✅ Se actualizaron {exitos} pasajeros con éxito.")
                                st.rerun()

                with st.expander("🚨 Zona de Peligro: Limpieza de Pasajeros"):
                    st.warning("Se borrarán todos los pasajeros de esta venta.")
                    confirm_pax = st.checkbox("Confirmar borrado de pasajeros", key="reset_pax_confirm")
                    if st.button("🗑️ Borrar Lista de Pasajeros", type="primary", disabled=not confirm_pax, use_container_width=True):
                        exito_p, msg_p = controller.borrar_pasajeros_venta(id_actual)
                        if exito_p: st.success(msg_p); st.rerun()
                        else: st.error(msg_p)
            else:
                st.info("Sin pasajeros registrados.")
        except Exception as e:
            st.error(f"Error cargando pasajeros: {e}")

    # Botón de Sincronización Unificado
    st.divider()
    if st.session_state.get('last_loaded_id_venta'):
        id_actual = st.session_state['last_loaded_id_venta']
        c_sync1, c_sync2, c_sync3 = st.columns([1, 2, 1])
        with c_sync2:
            if st.button("🔄 Sincronizar Itinerario", use_container_width=True, help="Aplica los cambios realizados en el constructor a esta venta."):
                from controllers.venta_controller import VentaController
                vc_temp = VentaController(supabase_client)
                exito_s, msg_s = vc_temp.sincronizar_venta_con_itinerario(id_actual)
                if exito_s:
                    st.success(msg_s)
                    st.rerun()
                else:
                    st.error(msg_s)

    # Botón de envío a contabilidad (Liquidación Final)
    st.divider()
    if st.session_state.get('last_loaded_id_venta'):
        id_actual = st.session_state['last_loaded_id_venta']
        
        c_arch_btn1, c_arch_btn2, c_arch_btn3 = st.columns([1, 2, 1])
        with c_arch_btn2:
            if st.button("📥 Liquidar y Archivar Expediente", type="primary", use_container_width=True, help="Finaliza el viaje, confirma todos los costos y oculta al pasajero de las listas activas."):
                exito, msg = controller.finalizar_liquidacion_venta(id_actual)
                if exito:
                    st.balloons()
                    st.success("✅ ¡Expediente Cerrado! El pasajero ha sido movido al historial y contabilidad.")
                    # Limpiar sesión para forzar refresco a una vista limpia
                    st.session_state['last_loaded_id_venta'] = None
                    st.session_state['simulador_adv_data'] = []
                    st.rerun()
                else:
                    st.error(msg)

        # Sección de Cancelación y Devoluciones (Matriz Contable)
        st.divider()
        with st.expander("❌ Módulo de Cancelaciones y Devoluciones (Matriz Contable)", expanded=False):
            st.markdown("### 📋 Cálculo Financiero por Cancelación de Reserva")
            st.caption("Esta herramienta calcula el reembolso sugerido al restar todos los costos incurridos y penalidades de los adelantos recaudados.")
            
            # Obtener datos consolidados para la cancelación
            res_c, err_c = controller.obtener_resumen_financiero_cancelacion(id_actual)
            
            if err_c:
                st.error(err_c)
            elif res_c:
                venta = res_c['venta']
                pasajeros = res_c['pasajeros']
                pagos = res_c['pagos']
                ingreso_recaudado = res_c['ingreso_recaudado']
                servicios = res_c['servicios']
                costo_incurrido_inicial = res_c['costo_incurrido_total']
                
                # ═══════════════════════════════════════════════════════════════
                # 1. INFORMACIÓN GENERAL
                # ═══════════════════════════════════════════════════════════════
                st.markdown("#### 👤 Información General")
                c1, c2 = st.columns(2)
                with c1:
                    st.write(f"**ID Venta:** {venta.get('id_venta')}")
                    st.write(f"**Cliente Principal:** {venta.get('cliente', {}).get('nombre') or 'No Registrado'}")
                    cliente_info = venta.get('cliente') or {}
                    lead_info = cliente_info.get('lead') or {}
                    celular = lead_info.get('numero_celular') or cliente_info.get('telefono') or 'No Registrado'
                    st.write(f"**Teléfono:** {celular}")
                with c2:
                    st.write(f"**Fecha Inicio:** {venta.get('fecha_inicio') or '---'}")
                    st.write(f"**Fecha Fin:** {venta.get('fecha_fin') or '---'}")
                    st.write(f"**Total Pax:** {venta.get('num_pasajeros') or len(pasajeros)} PAX")
                
                # ═══════════════════════════════════════════════════════════════
                # 2. DETALLE DE LIQUIDACIÓN Y COSTOS (Tabla)
                # ═══════════════════════════════════════════════════════════════
                st.markdown("#### 💰 Detalle de Liquidación y Costos")
                
                # Mantener en st.session_state las penalidades agregadas para esta venta
                session_key = f"penalidades_{id_actual}"
                if session_key not in st.session_state:
                    st.session_state[session_key] = []
                
                # Crear DataFrame con los servicios
                servs_data = []
                for s in servicios:
                    servs_data.append({
                        "Día": s['dia'],
                        "Proveedor Real": s['proveedor'],
                        "Tipo de Servicio": s['tipo_servicio'],
                        "Costo Unitario": f"{s['moneda']} {s['costo_unitario']:,.2f}",
                        "Pax/Cant": s['cantidad_pax'],
                        "TC": s['tc'],
                        "Total Soles (S/.)": s['total_soles']
                    })
                
                # Agregar las penalidades agregadas interactivamente
                for pen in st.session_state[session_key]:
                    servs_data.append({
                        "Día": "---",
                        "Proveedor Real": "⚠️ PENALIDAD PROVEEDOR",
                        "Tipo de Servicio": pen['descripcion'],
                        "Costo Unitario": f"{pen['moneda']} {pen['monto']:,.2f}",
                        "Pax/Cant": 1,
                        "TC": pen['tc'],
                        "Total Soles (S/.)": pen['total_soles']
                    })
                
                if servs_data:
                    df_servs = pd.DataFrame(servs_data)
                    st.dataframe(df_servs, use_container_width=True, hide_index=True)
                else:
                    st.info("No hay servicios registrados en la liquidación para calcular costos.")
                
                # Formulario para agregar Penalidad por Devolución (Inca Rail/Otros)
                st.markdown("##### ➕ Agregar Penalidad por Devolución (Proveedores)")
                col_p1, col_p2, col_p3, col_p4 = st.columns([2, 1, 1, 1])
                with col_p1:
                    desc_pen = st.text_input("Descripción de la Penalidad", value="PENALIDA POR DEVOLUCION", key=f"desc_pen_{id_actual}")
                with col_p2:
                    monto_pen = st.number_input("Monto", min_value=0.0, value=15.18, step=0.01, key=f"monto_pen_{id_actual}")
                with col_p3:
                    moneda_pen = st.selectbox("Moneda", ["USD", "PEN"], index=0, key=f"mon_pen_{id_actual}")
                with col_p4:
                    tc_pen = st.number_input("TC", min_value=1.0, value=3.38, step=0.01, key=f"tc_pen_{id_actual}")
                
                c_pbtn1, c_pbtn2 = st.columns([3, 1])
                with c_pbtn1:
                    if st.button("➕ Añadir Penalidad a Costos", use_container_width=True, key=f"btn_add_pen_{id_actual}"):
                        pen_total_soles = (monto_pen * tc_pen) if moneda_pen == "USD" else monto_pen
                        st.session_state[session_key].append({
                            "descripcion": desc_pen,
                            "monto": monto_pen,
                            "moneda": moneda_pen,
                            "tc": tc_pen,
                            "total_soles": pen_total_soles
                        })
                        st.toast("Penalidad agregada con éxito")
                        st.rerun()
                with c_pbtn2:
                    if st.session_state[session_key]:
                        if st.button("🧹 Limpiar", use_container_width=True, key=f"btn_clear_pen_{id_actual}"):
                            st.session_state[session_key] = []
                            st.toast("Penalidades limpiadas")
                            st.rerun()
                
                # --- NUEVO: Penalidad por Pasarela de Pago / Comisión de Tarjeta ---
                st.markdown("##### 💳 Penalidad por Pasarela de Pago (Comisión de Tarjeta / Banco)")
                monto_comision_pasarela = st.number_input(
                    "Comisión Pasarela (S/.)", 
                    min_value=0.0, 
                    value=float(ingreso_recaudado * 0.05), 
                    step=0.01, 
                    help="Monto de comisión cobrado por pasarelas (Visa, Niubiz, Culqi, etc.) sobre los adelantos recaudados. Sugerido: 5% del recaudado.",
                    key=f"gate_val_{id_actual}"
                )

                # Calcular Costo Total Final (Inicial + Penalidades de Proveedores + Penalidad de Pasarela)
                costo_penalidades_prov = sum(p['total_soles'] for p in st.session_state[session_key])
                costo_penalidades = costo_penalidades_prov + monto_comision_pasarela
                costo_total_final = costo_incurrido_inicial + costo_penalidades
                
                # ═══════════════════════════════════════════════════════════════
                # 3. RESUMEN FINANCIERO DE CANCELACIÓN (Dashboard)
                # ═══════════════════════════════════════════════════════════════
                st.markdown("#### 🏦 Resumen Financiero de Cancelación")
                
                col_res1, col_res2, col_res3 = st.columns(3)
                with col_res1:
                    st.metric("Total Recaudado (Adelantos)", f"S/. {ingreso_recaudado:,.2f}")
                with col_res2:
                    st.metric("Costo Incurrido Total", f"S/. {costo_total_final:,.2f}", delta=f"+S/. {costo_penalidades:,.2f}" if costo_penalidades > 0 else None, delta_color="inverse")
                
                diferencia_reembolso = ingreso_recaudado - costo_total_final
                with col_res3:
                    if diferencia_reembolso >= 0:
                        st.metric("Reembolso Máximo Sugerido", f"S/. {diferencia_reembolso:,.2f}", help="Monto máximo que puedes devolverle al cliente para quedar en punto de equilibrio (sin pérdidas).")
                    else:
                        st.metric("Riesgo de Pérdida Neta", f"S/. {abs(diferencia_reembolso):,.2f}", delta=f"-S/. {abs(diferencia_reembolso):,.2f}", delta_color="inverse", help="Alerta: Los costos incurridos superan los adelantos recibidos. Se requiere cobrar penalidad adicional.")
                
                # Mostrar el desglose de pagos registrados
                st.markdown("##### Historial de Pagos Recibidos")
                pagos_data = []
                for p in pagos:
                    pagos_data.append({
                        "ID": p['id_pago'],
                        "Fecha": p['fecha_pago'],
                        "Tipo": p['tipo_pago'],
                        "Monto Original": f"{p['moneda']} {p['monto']:,.2f}",
                        "TC": p['tc'],
                        "Total Soles (S/.)": f"S/. {p['monto_soles']:,.2f}"
                    })
                if pagos_data:
                    st.dataframe(pd.DataFrame(pagos_data), use_container_width=True, hide_index=True)
                else:
                    st.warning("No hay pagos registrados para este viaje.")
                
                # ═══════════════════════════════════════════════════════════════
                # 4. FORMULARIO DE DEVOLUCIÓN Y CONFIRMACIÓN
                # ═══════════════════════════════════════════════════════════════
                st.divider()
                st.markdown("#### 🔒 Formulario de Confirmación y Devolución")
                
                c_form1, c_form2 = st.columns(2)
                with c_form1:
                    monto_a_reembolsar = st.number_input("Monto Real a Reembolsar al Cliente (S/.)", min_value=0.0, max_value=max(0.0, float(diferencia_reembolso)), value=max(0.0, float(diferencia_reembolso)), step=1.0, key=f"refund_val_{id_actual}")
                with c_form2:
                    metodo_reembolso = st.selectbox("Método de Reembolso", ["TRANSFERENCIA", "EFECTIVO", "YAPE", "PLIN", "INTERBANK", "VISA", "OTRO"], index=0, key=f"refund_method_{id_actual}")
                
                obs_cancelacion = st.text_area("Observaciones de la Cancelación (Motivos y Acuerdos)", value="Pasajero cancela por motivos de salud / fuerza mayor.", placeholder="Escribe aquí los detalles del acuerdo...", key=f"refund_obs_{id_actual}")
                
                st.warning("⚠️ **ATENCIÓN:** Esta acción marcará el viaje como CANCELADO en el sistema de manera definitiva, desactivando todos los servicios en la lista activa de operaciones y registrando el egreso correspondiente por reembolso. Esta acción no se puede deshacer.")
                
                confirmar_check = st.checkbox("Entiendo la acción y confirmo que los datos financieros cuadran perfectamente.", key=f"refund_confirm_{id_actual}")
                
                if st.button("❌ Procesar Cancelación Definitiva de Reserva", type="primary", disabled=not confirmar_check, use_container_width=True, key=f"btn_execute_cancel_{id_actual}"):
                    with st.spinner("Procesando cancelación física y contable..."):
                        # Construir desglose financiero para guardarlo en la Base de Datos en el log de pagos
                        desglose_obs = (
                            f"{obs_cancelacion}\n\n"
                            f"--- DESGLOSE CONTABLE DE CANCELACIÓN ---\n"
                            f"• Total Recaudado (Adelantos): S/. {ingreso_recaudado:,.2f}\n"
                            f"• Costo Incurrido en Servicios: S/. {costo_incurrido_inicial:,.2f}\n"
                            f"• Penalidad Pasarela: S/. {monto_comision_pasarela:,.2f}\n"
                            f"• Penalidades de Proveedores: S/. {costo_penalidades_prov:,.2f}\n"
                            f"• Total Retenido (Penalidades): S/. {costo_penalidades:,.2f}\n"
                            f"• Monto Real Reembolsado: S/. {monto_a_reembolsar:,.2f}"
                        )
                        
                        exito, msg = controller.ejecutar_cancelacion_reserva(
                            id_venta=id_actual,
                            costo_penalidad=costo_penalidades,
                            descripcion_penalidad="Penalidad acumulada en cancelación",
                            monto_reembolsado=monto_a_reembolsar,
                            metodo_reembolso=metodo_reembolso,
                            observaciones=desglose_obs
                        )
                        if exito:
                            # Limpiar sesión y variables de penalidades
                            st.session_state[session_key] = []
                            st.session_state['last_loaded_id_venta'] = None
                            st.session_state['simulador_adv_data'] = []
                            st.success("✅ ¡Viaje Cancelado con éxito! Todos los estados han sido actualizados y el reembolso contable ha sido ingresado.")
                            st.balloons()
                            st.rerun()
                        else:
                            st.error(msg)

        # Cancelación parcial (solo algunos pasajeros del paquete)
        st.divider()
        with st.expander("🟠 Cancelación Parcial por Pasajero (Matriz Contable)", expanded=False):
            render_modulo_cancelacion_parcial(controller, id_actual)

def render_modulo_cancelacion_parcial(controller, id_venta):
    """
    Matriz de cancelación cuando solo parte del grupo cancela (ej. 2 de 10 pax).
    No cancela la venta completa ni los servicios del resto del grupo.
    """
    st.markdown("### 👥 Cancelación parcial del grupo")
    st.caption(
        "Selecciona los pasajeros que se dan de baja. El viaje sigue activo para los demás. "
        "Los montos se prorratean según cantidad de pax cancelados vs total de la venta."
    )

    res_base, err_base = controller.obtener_resumen_financiero_cancelacion(id_venta)
    if err_base:
        st.error(err_base)
        return
    if not res_base:
        st.warning("No hay datos de la venta.")
        return

    todos = res_base.get('pasajeros') or []
    activos = [p for p in todos if str(p.get('estado_pasajero') or 'ACTIVO').upper() != 'CANCELADO']
    ya_cancelados = [p for p in todos if str(p.get('estado_pasajero') or '').upper() == 'CANCELADO']

    if ya_cancelados:
        with st.expander(f"Ver {len(ya_cancelados)} pasajero(s) ya cancelados en esta venta", expanded=False):
            for p in ya_cancelados:
                st.write(f"• {p.get('nombre_completo', '---')} — {p.get('fecha_cancelacion', '')[:10] if p.get('fecha_cancelacion') else 'sin fecha'}")

    if not activos:
        st.info("No hay pasajeros activos en esta venta. Todos fueron cancelados o no hay rooming cargado.")
        return

    opciones = {
        f"{p.get('nombre_completo', 'Sin nombre')} | Doc: {p.get('numero_documento') or '---'}": int(p['id_pasajero'])
        for p in activos
    }
    labels_sel = st.multiselect(
        "Pasajeros a cancelar",
        options=list(opciones.keys()),
        help="Elige solo quienes se retiran del paquete.",
        key=f"pax_cancel_sel_{id_venta}",
    )
    ids_sel = [opciones[l] for l in labels_sel]

    if not ids_sel:
        st.info("Selecciona uno o más pasajeros para ver la matriz financiera.")
        return

    res_p, err_p = controller.obtener_resumen_cancelacion_parcial(id_venta, ids_sel)
    if err_p:
        st.error(err_p)
        return

    venta = res_p['venta']
    st.markdown("#### 👤 Información General (Baja de Pasajeros)")
    c1, c2 = st.columns(2)
    with c1:
        st.write(f"**ID Venta:** {venta.get('id_venta')}")
        st.write(f"**Cliente Principal:** {venta.get('cliente', {}).get('nombre') or 'No Registrado'}")
        cliente_info = venta.get('cliente') or {}
        lead_info = cliente_info.get('lead') or {}
        celular = lead_info.get('numero_celular') or cliente_info.get('telefono') or 'No Registrado'
        st.write(f"**Teléfono:** {celular}")
    with c2:
        st.write(f"**Pasajeros a Cancelar:** {res_p['n_cancelar']} PAX")
        st.write(f"**Pasajeros que Continúan:** {len(res_p['pasajeros_activos_restantes'])} PAX")
        st.write(f"**Prorrateo Aplicado:** {res_p['n_cancelar']} / {res_p['n_total_ref']} = **{res_p['factor']*100:.1f}%**")

    session_key = f"penalidades_parcial_{id_venta}"
    if session_key not in st.session_state:
        st.session_state[session_key] = []

    moneda_v = res_p.get('moneda_venta', 'USD')
    if moneda_v == "USD":
        sym = "$"
    elif moneda_v == "PEN":
        sym = "S/."
    elif moneda_v == "EUR":
        sym = "€"
    else:
        sym = moneda_v

    factor = res_p['factor']
    factor_pct = factor * 100.0

    st.markdown(f"#### 💰 Detalle de Liquidación y Costos (Prorrateado al {factor_pct:.1f}%)")
    
    # Crear DataFrame con los servicios prorrateados
    servs_data = []
    for s in res_p.get('servicios', []):
        c_unit = float(s.get('costo_unitario') or 0.0)
        pax = float(s.get('cantidad_pax') or 1.0)
        tc = float(s.get('tc') or 3.80)
        moneda_s = s.get('moneda') or 'USD'
        total_soles = s.get('total_soles') or 0.0
        
        # Costo prorrateado en la moneda de la venta
        moneda_s_up = moneda_s.upper()
        if moneda_s_up == moneda_v:
            costo_prorrateado_mv = (c_unit * pax) * factor
        elif moneda_s_up == 'USD' and moneda_v == 'PEN':
            costo_prorrateado_mv = (c_unit * pax * tc) * factor
        elif moneda_s_up == 'PEN' and moneda_v == 'USD':
            costo_prorrateado_mv = ((c_unit * pax) / tc) * factor if tc > 0 else ((c_unit * pax) / 3.80) * factor
        else:
            costo_prorrateado_mv = (total_soles * factor) / res_p.get('tc_venta', 3.80) if moneda_v == 'USD' else (total_soles * factor)

        servs_data.append({
            "Día": s['dia'],
            "Proveedor Real": s['proveedor'],
            "Tipo de Servicio": s['tipo_servicio'],
            "Costo Unitario": f"{moneda_s} {c_unit:,.2f}",
            "Pax Original": pax,
            "Total Soles Original": f"S/. {total_soles:,.2f}",
            f"Prorrateado ({moneda_v})": f"{sym} {costo_prorrateado_mv:,.2f}"
        })

    # Agregar penalidades interactivas
    for pen in st.session_state[session_key]:
        servs_data.append({
            "Día": "---",
            "Proveedor Real": "⚠️ PENALIDAD PROVEEDOR",
            "Tipo de Servicio": pen['descripcion'],
            "Costo Unitario": f"{pen['moneda']} {pen['monto']:,.2f}",
            "Pax Original": 1,
            "Total Soles Original": "---",
            f"Prorrateado ({moneda_v})": f"{sym} {pen['total_moneda_venta']:,.2f}"
        })

    if servs_data:
        df_servs = pd.DataFrame(servs_data)
        st.dataframe(df_servs, use_container_width=True, hide_index=True)
    else:
        st.info("No hay servicios registrados en la liquidación para calcular costos.")

    # Formulario para agregar Penalidad
    st.markdown("##### ➕ Agregar Penalidad por Devolución (Proveedores)")
    col_p1, col_p2, col_p3, col_p4 = st.columns([2, 1, 1, 1])
    with col_p1:
        desc_pen = st.text_input("Descripción de la Penalidad", value="PENALIDAD POR DEVOLUCION", key=f"desc_pen_p_{id_venta}")
    with col_p2:
        monto_pen = st.number_input("Monto", min_value=0.0, value=0.0, step=0.01, key=f"monto_pen_p_{id_venta}")
    with col_p3:
        moneda_pen = st.selectbox("Moneda", ["USD", "PEN"], index=0 if moneda_v=="USD" else 1, key=f"mon_pen_p_{id_venta}")
    with col_p4:
        tc_pen = st.number_input("TC", min_value=1.0, value=float(res_p.get('tc_venta') or 3.8), step=0.01, key=f"tc_pen_p_{id_venta}")

    c_pbtn1, c_pbtn2 = st.columns([3, 1])
    with c_pbtn1:
        if st.button("➕ Añadir Penalidad a Costos", use_container_width=True, key=f"btn_add_pen_p_{id_venta}"):
            if moneda_v == 'USD':
                total_mv = monto_pen if moneda_pen == "USD" else (monto_pen / tc_pen)
            else:
                total_mv = (monto_pen * tc_pen) if moneda_pen == "USD" else monto_pen
                
            st.session_state[session_key].append({
                "descripcion": desc_pen, "monto": monto_pen, "moneda": moneda_pen, "tc": tc_pen, "total_moneda_venta": total_mv
            })
            st.toast("Penalidad agregada con éxito")
            st.rerun()
    with c_pbtn2:
        if st.session_state[session_key]:
            if st.button("🧹 Limpiar", use_container_width=True, key=f"btn_clear_pen_p_{id_venta}"):
                st.session_state[session_key] = []
                st.toast("Penalidades limpiadas")
                st.rerun()

    # Comisión pasarela
    st.markdown("##### 💳 Penalidad por Pasarela de Pago (Comisión de Tarjeta / Banco)")
    ingreso_prorr = float(res_p['ingreso_prorrateado'])
    monto_comision_pasarela = st.number_input(
        f"Comisión Pasarela ({moneda_v})",
        min_value=0.0,
        value=float(ingreso_prorr * 0.05),
        step=0.01,
        help="Monto de comisión cobrado por pasarelas (Visa, Niubiz, Culqi, etc.) sobre los adelantos recaudados. Sugerido: 5% del recaudado.",
        key=f"gate_val_p_{id_venta}"
    )

    costo_prorr = float(res_p['costo_prorrateado'])
    pen_prov = sum(p['total_moneda_venta'] for p in st.session_state[session_key])
    penalidades = pen_prov + monto_comision_pasarela
    costo_total_parcial = costo_prorr + penalidades
    reembolso_sugerido = ingreso_prorr - costo_total_parcial

    # Resumen Financiero
    st.markdown(f"#### 🏦 Resumen Financiero de Cancelación Parcial (en {moneda_v})")
    col_res1, col_res2, col_res3 = st.columns(3)
    with col_res1:
        st.metric("Adelanto Recaudado Prorrateado", f"{sym} {ingreso_prorr:,.2f}")
    with col_res2:
        st.metric("Costo Incurrido Prorrateado", f"{sym} {costo_total_parcial:,.2f}", delta=f"+{sym} {penalidades:,.2f}" if penalidades > 0 else None, delta_color="inverse")
    with col_res3:
        if reembolso_sugerido >= 0:
            st.metric("Reembolso Sugerido", f"{sym} {reembolso_sugerido:,.2f}", help="Monto sugerido a devolver para quedar en equilibrio para estos pasajeros.")
        else:
            st.metric("Riesgo de Pérdida Neta (Parcial)", f"{sym} {abs(reembolso_sugerido):,.2f}", delta=f"-{sym} {abs(reembolso_sugerido):,.2f}", delta_color="inverse", help="Alerta: Los costos prorrateados superan los adelantos.")

    # Historial de Pagos Recibidos
    st.markdown(f"##### Historial de Pagos Recibidos (Prorrateado al {factor_pct:.1f}%)")
    pagos_data = []
    for p in res_p.get('pagos', []):
        monto_orig = p['monto']
        monto_soles = p['monto_soles']
        
        if moneda_v == 'USD':
            monto_mv = monto_soles / res_p.get('tc_venta', 3.80)
        else:
            monto_mv = monto_soles
            
        monto_prorr_mv = monto_mv * factor
        
        pagos_data.append({
            "ID": p['id_pago'],
            "Fecha": p['fecha_pago'],
            "Tipo": p['tipo_pago'],
            "Monto Original": f"{p['moneda']} {monto_orig:,.2f}",
            "TC Pago": p['tc'],
            f"Prorrateado ({moneda_v})": f"{sym} {monto_prorr_mv:,.2f}"
        })

    if pagos_data:
        st.dataframe(pd.DataFrame(pagos_data), use_container_width=True, hide_index=True)
    else:
        st.warning("No hay pagos registrados para este viaje.")

    st.markdown("##### Pasajeros que se cancelan")
    for p in res_p['pasajeros_seleccionados']:
        st.write(f"• {p.get('nombre_completo', '---')}")

    st.divider()
    st.markdown("#### 🔒 Formulario de Confirmación y Devolución Parcial")
    
    c_form1, c_form2 = st.columns(2)
    with c_form1:
        monto_reemb = st.number_input(
            f"Monto Real a Reembolsar al Cliente ({sym})",
            min_value=0.0,
            value=max(0.0, float(reembolso_sugerido)),
            step=1.0,
            key=f"refund_parcial_{id_venta}",
        )
    with c_form2:
        metodo = st.selectbox(
            "Método de Reembolso",
            ["TRANSFERENCIA", "EFECTIVO", "YAPE", "PLIN", "INTERBANK", "VISA", "OTRO"],
            key=f"method_parcial_{id_venta}",
        )

    obs = st.text_area(
        "Observaciones de la Cancelación Parcial (Motivos y Acuerdos)", 
        value="Pasajero cancela por motivos de salud / fuerza mayor.",
        placeholder="Escribe aquí los detalles del acuerdo...",
        key=f"obs_parcial_{id_venta}"
    )
    
    ajustar_pax = st.checkbox(
        "Reducir cantidad en calendario operativo (venta_tour)",
        value=True,
        help="Resta los pax cancelados del campo cantidad en cada día del itinerario.",
        key=f"adj_vt_p_{id_venta}",
    )
    st.warning(
        "⚠️ **ATENCIÓN:** Esta acción marcará como CANCELADO únicamente a los pasajeros seleccionados. "
        "El resto del grupo sigue activo. Se registrará la salida del egreso contable en la moneda de la venta."
    )
    confirm = st.checkbox("Entiendo la acción y confirmo que los datos financieros cuadran perfectamente.", key=f"confirm_parcial_{id_venta}")

    if st.button(
        "🟠 Procesar Cancelación Parcial de Reserva",
        type="primary",
        disabled=not confirm,
        use_container_width=True,
        key=f"btn_parcial_{id_venta}",
    ):
        desglose = (
            f"{obs}\n\n"
            f"--- DESGLOSE CONTABLE DE CANCELACIÓN PARCIAL ---\n"
            f"• Pasajeros de baja: {', '.join([p.get('nombre_completo','') for p in res_p['pasajeros_seleccionados']])}\n"
            f"• Prorrateo aplicado: {res_p['n_cancelar']}/{res_p['n_total_ref']} ({factor_pct:.1f}%)\n"
            f"• Adelanto Recaudado Prorrateado: {sym} {ingreso_prorr:,.2f}\n"
            f"• Costo Incurrido Prorrateado: {sym} {costo_prorr:,.2f}\n"
            f"• Penalidad Pasarela: {sym} {monto_comision_pasarela:,.2f}\n"
            f"• Penalidades de Proveedores: {sym} {pen_prov:,.2f}\n"
            f"• Total Retenido (Penalidades): {sym} {penalidades:,.2f}\n"
            f"• Monto Real Reembolsado: {sym} {monto_reemb:,.2f}"
        )
        with st.spinner("Registrando baja parcial..."):
            ok, msg = controller.ejecutar_cancelacion_parcial(
                id_venta=id_venta,
                ids_pasajeros=ids_sel,
                monto_reembolsado=monto_reemb,
                metodo_reembolso=metodo,
                observaciones=desglose,
                penalidad_total=penalidades,
                ajustar_cantidad_itinerario=ajustar_pax,
            )
        if ok:
            st.session_state[session_key] = []
            st.success("✅ " + msg)
            st.rerun()
        else:
            st.error(msg)

def render_directorio_proveedores(supabase_client):
    """Módulo profesional para la gestión de proveedores con soporte JSONB dinámico."""
    from controllers.proveedor_controller import ProveedorController
    import json
    prov_ctrl = ProveedorController(supabase_client)

    st.subheader("🏢 Gestión de Socios Estratégicos (Proveedores)", divider="red")
    
    # ═══════════════════════════════════════════════════════════════
    # 1. GESTIÓN DE ESTADO (Draft)
    # ═══════════════════════════════════════════════════════════════
    if 'prov_edit_id' not in st.session_state:
        st.session_state.prov_edit_id = None
    if 'prov_draft' not in st.session_state:
        st.session_state.prov_draft = {}

    # ═══════════════════════════════════════════════════════════════
    # 2. SELECTOR DE PROVEEDOR
    # ═══════════════════════════════════════════════════════════════
    listado_prov = prov_ctrl.obtener_proveedores()
    mapa_nombres = {p['nombre_comercial']: p for p in listado_prov}
    
    col_sel1, col_sel2 = st.columns([2, 1])
    prov_sel = col_sel1.selectbox(
        "🔍 Buscar o Seleccionar Proveedor:", 
        ["--- Nuevo Proveedor ---"] + list(mapa_nombres.keys()),
        index=0 if st.session_state.prov_edit_id is None else (list(mapa_nombres.keys()).index(next(k for k, v in mapa_nombres.items() if v['id_proveedor'] == st.session_state.prov_edit_id)) + 1 if st.session_state.prov_edit_id in [p['id_proveedor'] for p in listado_prov] else 0)
    )

    # Cargar datos al cambiar selección
    if prov_sel == "--- Nuevo Proveedor ---":
        if st.session_state.prov_edit_id is not None:
            st.session_state.prov_edit_id = None
            st.session_state.prov_draft = {
                "nombre_comercial": "", "ruc": "", "email": "", "persona_contacto": "",
                "contacto_telefono": "", "pais": "Perú", "url_drive": "",
                "servicios_ofrecidos": ["GUIADO"], "cuentas_bancarias": [],
                "puntos_operacion": [], "detalles_categoria": {}, "activo": True
            }
            st.rerun()
    else:
        p_data = mapa_nombres[prov_sel]
        if st.session_state.prov_edit_id != p_data['id_proveedor']:
            st.session_state.prov_edit_id = p_data['id_proveedor']
            # Cargar todo el objeto a memoria (Draft)
            st.session_state.prov_draft = p_data.copy()
            # Asegurar que los JSON no sean None
            for key in ['cuentas_bancarias', 'puntos_operacion']:
                if not st.session_state.prov_draft.get(key): st.session_state.prov_draft[key] = []
            if not st.session_state.prov_draft.get('detalles_categoria'): 
                st.session_state.prov_draft['detalles_categoria'] = {}
            st.rerun()

    # Si no hay draft inicializado (caso primer carga), inicializarlo vacío
    if not st.session_state.prov_draft:
        st.session_state.prov_draft = {
            "nombre_comercial": "", "ruc": "", "email": "", "persona_contacto": "",
            "contacto_telefono": "", "pais": "Perú", "url_drive": "",
            "servicios_ofrecidos": ["GUIADO"], "cuentas_bancarias": [],
            "puntos_operacion": [], "detalles_categoria": {}, "activo": True
        }

    draft = st.session_state.prov_draft

    # ═══════════════════════════════════════════════════════════════
    # 3. FORMULARIO DINÁMICO
    # ═══════════════════════════════════════════════════════════════
    st.write("---")
    
    # --- BLOQUE A: INFORMACIÓN BÁSICA ---
    with st.expander("👤 Información General y Contacto", expanded=True):
        c1, c2, c3 = st.columns([2, 1, 1])
        draft['nombre_comercial'] = c1.text_input("Nombre / Razón Social*", value=draft.get('nombre_comercial', ''))
        draft['ruc'] = c2.text_input("RUC / Tax ID", value=draft.get('ruc', ''))
        draft['persona_contacto'] = c3.text_input("Contacto Principal", value=draft.get('persona_contacto', ''))
        
        ca, cb, cc = st.columns(3)
        draft['contacto_telefono'] = ca.text_input("Teléfono / WhatsApp", value=draft.get('contacto_telefono', ''))
        draft['email'] = cb.text_input("Email de Reservas", value=draft.get('email', ''))
        draft['url_drive'] = cc.text_input("🔗 Link Drive (Tarifarios/Docs)", value=draft.get('url_drive', ''))

        serv_base = ["GUIA", "TRANSPORTE", "ALOJAMIENTO", "ALIMENTACION", "TICKETS", "AGENCIA", "OPERADOR", "OTROS"]
        draft['servicios_ofrecidos'] = st.multiselect(
            "Servicios que brinda*", 
            options=list(set(serv_base + (draft.get('servicios_ofrecidos') or []))),
            default=draft.get('servicios_ofrecidos', ["GUIA"])
        )

    # --- BLOQUE B: CUENTAS BANCARIAS (DINÁMICO) ---
    with st.expander(f"💳 Información de Pagos ({len(draft['cuentas_bancarias'])} cuentas)", expanded=False):
        for idx, cuenta in enumerate(draft['cuentas_bancarias']):
            colb1, colb2, colb3, colb4, colb_del = st.columns([1, 1, 1.5, 2, 0.5])
            cuenta['banco'] = colb1.text_input(f"Banco", value=cuenta.get('banco', ''), key=f"bnk_{idx}")
            cuenta['moneda'] = colb2.selectbox(f"Moneda", ["PEN", "USD"], index=0 if cuenta.get('moneda')=="PEN" else 1, key=f"mon_{idx}")
            cuenta['nro'] = colb3.text_input(f"N° Cuenta", value=cuenta.get('nro', ''), key=f"nro_{idx}")
            cuenta['cci'] = colb4.text_input(f"CCI (20 dígitos)", value=cuenta.get('cci', ''), key=f"cci_{idx}")
            if colb_del.button("❌", key=f"del_bnk_{idx}"):
                draft['cuentas_bancarias'].pop(idx)
                st.rerun()
        
        if st.button("➕ Agregar Cuenta Bancaria", use_container_width=True):
            draft['cuentas_bancarias'].append({"banco": "", "moneda": "PEN", "nro": ""})
            st.rerun()

    # --- BLOQUE C: LOGÍSTICA Y OPERACIÓN ---
    with st.expander("📍 Zonas de Operación y Logística", expanded=False):
        c_zonas_1, c_zonas_2 = st.columns([3, 1])
        nueva_zona = c_zonas_1.text_input("Agregar zona de operación:", placeholder="Ej: Cusco Centro, Ollantaytambo, Lima...")
        if c_zonas_2.button("➕ Añadir", use_container_width=True) and nueva_zona:
            if nueva_zona not in draft['puntos_operacion']:
                draft['puntos_operacion'].append(nueva_zona)
                st.rerun()
        
        if draft['puntos_operacion']:
            st.write("Zonas registradas:")
            cols_z = st.columns(4)
            for z_idx, z_val in enumerate(draft['puntos_operacion']):
                with cols_z[z_idx % 4]:
                    if st.button(f"{z_val} ❌", key=f"z_{z_idx}"):
                        draft['puntos_operacion'].pop(z_idx)
                        st.rerun()

    # --- BLOQUE D: CAMPOS INTELIGENTES POR CATEGORÍA ---
    # Detectar categoría principal para mostrar campos específicos
    servs = draft.get('servicios_ofrecidos', [])
    detalles = draft['detalles_categoria']
    
    if any(s in servs for s in ["GUIA", "GUIADO"]):
        with st.expander("🎓 Detalles Especializados: GUÍA", expanded=True):
            cg1, cg2 = st.columns(2)
            detalles['idiomas'] = cg1.text_input("Idiomas que habla", value=detalles.get('idiomas', 'Español, Inglés'))
            detalles['nro_carnet'] = cg2.text_input("N° Carnet GRL / Oficial", value=detalles.get('nro_carnet', ''))
            detalles['especialidad'] = st.text_input("Especialidad (Aventura, Cultural, etc)", value=detalles.get('especialidad', ''))

    if any(s in servs for s in ["ALOJAMIENTO", "HOTEL"]):
        with st.expander("🏨 Detalles Especializados: HOTEL", expanded=True):
            ch1, ch2, ch3 = st.columns(3)
            detalles['estrellas'] = ch1.selectbox("Categoría", ["1*", "2*", "3*", "4*", "5*", "Boutique", "Hostal"], index=2)
            detalles['check_in'] = ch2.text_input("Hora Check-In", value=detalles.get('check_in', '12:00 PM'))
            detalles['desayuno'] = ch3.toggle("¿Incluye Desayuno?", value=detalles.get('desayuno', True))

    if any(s in servs for s in ["TRANSPORTE"]):
        with st.expander("🚐 Detalles Especializados: TRANSPORTE", expanded=True):
            ct1, ct2, ct3 = st.columns(3)
            detalles['vehiculo'] = ct1.text_input("Marca/Modelo", value=detalles.get('vehiculo', ''))
            detalles['placa'] = ct2.text_input("Placa", value=detalles.get('placa', ''))
            detalles['capacidad'] = ct3.number_input("Capacidad Pax", value=detalles.get('capacidad', 1), min_value=1)

    # ═══════════════════════════════════════════════════════════════
    # 4. ACCIONES DE GUARDADO
    # ═══════════════════════════════════════════════════════════════
    st.write("---")
    draft['activo'] = st.toggle("Proveedor Activo", value=draft.get('activo', True))
    
    col_acc1, col_acc2 = st.columns(2)
    
    if col_acc1.button("💾 GUARDAR CAMBIOS", type="primary", use_container_width=True):
        if not draft['nombre_comercial']:
            st.error("Error: El nombre es obligatorio.")
        else:
            if st.session_state.prov_edit_id:
                # MODO EDICION
                exito, msg = prov_ctrl.actualizar_proveedor(
                    st.session_state.prov_edit_id,
                    draft['nombre_comercial'], draft['servicios_ofrecidos'],
                    draft['contacto_telefono'], draft.get('pais', 'Perú'),
                    draft['activo'], draft['ruc'], draft['email'],
                    draft['persona_contacto'], draft['url_drive'],
                    draft['cuentas_bancarias'], draft['puntos_operacion'],
                    draft['detalles_categoria']
                )
            else:
                # MODO REGISTRO
                exito, msg = prov_ctrl.registrar_proveedor(
                    draft['nombre_comercial'], draft['servicios_ofrecidos'],
                    draft['contacto_telefono'], draft.get('pais', 'Perú'),
                    draft['ruc'], draft['email'], draft['persona_contacto'],
                    draft['url_drive'], draft['cuentas_bancarias'],
                    draft['puntos_operacion'], draft['detalles_categoria']
                )
            
            if exito:
                st.success(msg)
                st.balloons()
                # Limpiar estado
                st.session_state.prov_edit_id = None
                st.session_state.prov_draft = {}
                st.rerun()
            else:
                st.error(msg)

    if col_acc2.button("🧹 LIMPIAR / CANCELAR", use_container_width=True):
        st.session_state.prov_edit_id = None
        st.session_state.prov_draft = {}
        st.rerun()

    # ═══════════════════════════════════════════════════════════════
    # 5. LISTADO GENERAL
    # ═══════════════════════════════════════════════════════════════
    with st.expander("📜 Ver Directorio General Completo", expanded=False):
        if not listado_prov:
            st.info("Sin proveedores registrados.")
        else:
            df_view = pd.DataFrame(listado_prov)
            # Limpiar servicios para vista
            df_view['Servicios'] = df_view['servicios_ofrecidos'].apply(lambda x: ", ".join(x) if isinstance(x, list) else x)
            st.dataframe(
                df_view,
                column_order=["nombre_comercial", "Servicios", "contacto_telefono", "ruc", "email", "activo"],
                column_config={
                    "nombre_comercial": "Proveedor",
                    "contacto_telefono": "Teléfono",
                    "ruc": "RUC",
                    "email": "Email",
                    "activo": st.column_config.CheckboxColumn("Activo")
                },
                use_container_width=True,
                hide_index=True
            )


    # ═══════════════════════════════════════════════════════════════
    # 3. TABLA DE PROVEEDORES EXISTENTES
    # ═══════════════════════════════════════════════════════════════
    st.write("### 📜 Directorio General")
    proveedores = listado_prov # Usamos la misma lista ya cargada
    
    if not proveedores:
        st.info("Aún no hay proveedores registrados en el directorio.")
    else:
        df_prov = pd.DataFrame(proveedores)
        
        # Formatear la columna de servicios para que sea más legible si es una lista
        if 'servicios_ofrecidos' in df_prov.columns:
            df_prov['servicios_ofrecidos'] = df_prov['servicios_ofrecidos'].apply(lambda x: ", ".join(x) if isinstance(x, list) else x)
        
        st.dataframe(
            df_prov,
            column_order=["nombre_comercial", "servicios_ofrecidos", "contacto_telefono", "pais", "activo"],
            column_config={
                "nombre_comercial": "Proveedor",
                "servicios_ofrecidos": "Servicios",
                "contacto_telefono": "Teléfono",
                "pais": "País",
                "activo": st.column_config.CheckboxColumn("Activo")
            },
            hide_index=True,
            use_container_width=True
        )

        st.write("---")
        with st.expander("🗑️ Eliminar Proveedor", expanded=False):
            st.warning("⚠️ Cuidado: Eliminar un proveedor es una acción irreversible.")
            opciones_eliminar = {f"{p['nombre_comercial']} (RUC: {p.get('ruc', 'N/A')})": p['id_proveedor'] for p in proveedores}
            prov_a_eliminar = st.selectbox("Seleccione un proveedor para eliminar:", ["---"] + list(opciones_eliminar.keys()), key="del_prov_sel")
            
            if prov_a_eliminar != "---":
                if st.button("🚨 Confirmar Eliminación", type="primary", key="btn_del_prov"):
                    id_eliminar = opciones_eliminar[prov_a_eliminar]
                    exito, msg = prov_ctrl.eliminar_proveedor(id_eliminar)
                    if exito:
                        st.success(msg)
                        st.rerun()
                    else:
                        st.error(msg)
