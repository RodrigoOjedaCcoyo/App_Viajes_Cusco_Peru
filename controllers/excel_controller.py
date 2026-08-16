# controllers/excel_controller.py
import pandas as pd
from io import BytesIO
import datetime
from openpyxl.styles import Alignment, Font, PatternFill

class ExcelController:
    """Controlador para la generación de documentos Excel a partir de datos del itinerario."""

    def generar_resumen_itinerario_xlsx(self, datos_render: dict, precios_reales: dict = None, nombre_cliente: str = None, num_pax: int = None) -> BytesIO:
        """
        Genera un archivo XLSX resumido para operaciones basado estrictamente en el itinerario.
        Permite inyectar 'precios_reales' (mapa {n_linea: precio}) desde la base de datos.
        """
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "RESUMEN_OPERATIVO"
        
        # --- ESTILOS ---
        bold_f = Font(bold=True)
        center_al = Alignment(horizontal='center', vertical='center')
        top_al = Alignment(vertical='top', wrap_text=True)
        
        # --- HEADER ---
        ws.merge_cells('A1:E1')
        titulo_itin = (datos_render.get('titulo') or datos_render.get('paquete_nombre') or 'ITINERARIO').upper()
        ws['A1'] = f"RESUMEN OPERATIVO: {titulo_itin}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = center_al
        
        ws['A3'] = "FECHA INICIO:"
        
        # --- EXTRACCIÓN ROBUSTA DE FECHA (Misma lógica que Ventas) ---
        f_inicio = datos_render.get('fecha_viaje') or datos_render.get('fecha_inicio') or datos_render.get('fechaViaje') or datos_render.get('fecha')
        
        ci = datos_render.get('control_interno', {})
        if not f_inicio and ci:
            f_inicio = ci.get('fecha_inicio') or ci.get('fecha_llegada') or ci.get('fecha_viaje')
            
        if not f_inicio:
            dias_itin = datos_render.get('itinerario') or datos_render.get('days') or datos_render.get('itinerario_detalles')
            if dias_itin and isinstance(dias_itin, list) and len(dias_itin) > 0 and isinstance(dias_itin[0], dict):
                f_inicio = dias_itin[0].get('fecha')
        
        f_inicio_str = str(f_inicio).strip() if f_inicio else "---"
        
        # Extracción de fecha fin robusta
        f_fin_str = datos_render.get('fecha_fin') or datos_render.get('fechaFin')
        if not f_fin_str and f_inicio and f_inicio_str != "---":
            # Calcular base a duración
            duracion_raw = datos_render.get('duracion')
            if duracion_raw and isinstance(duracion_raw, str) and 'D' in duracion_raw.upper():
                try:
                    num_dias_str = ''.join(filter(str.isdigit, duracion_raw.split('D')[0]))
                    if num_dias_str:
                        num_dias = int(num_dias_str)
                        # Intento de parseo
                        f_clean = f_inicio_str.replace(" ", "")
                        obj_f_inicio = None
                        if '/' in f_clean:
                            for fmt in ("%d/%m/%Y", "%Y/%m/%d", "%m/%d/%Y"):
                                try:
                                    obj_f_inicio = datetime.datetime.strptime(f_clean, fmt).date()
                                    break
                                except: pass
                        elif '-' in f_clean:
                            obj_f_inicio = datetime.date.fromisoformat(f_clean[:10])
                        
                        if obj_f_inicio:
                            f_fin_obj = obj_f_inicio + datetime.timedelta(days=num_dias - 1)
                            f_fin_str = f_fin_obj.strftime("%d/%m/%Y")
                except Exception:
                    pass
        
        if not f_fin_str:
            f_fin_str = "---"

        ws['B3'] = f_inicio_str
        ws['D3'] = "FECHA FINAL:"
        ws['E3'] = f_fin_str
        
        ws['A4'] = "PASAJERO:"
        # Mapeo robusto de nombre: PRIORIDAD PARÁMETRO > POSIBLES_NOMBRES > PLACEHOLDER
        control_int = datos_render.get("control_interno") or {}
        posibles_nombres = [
            nombre_cliente, # Parámetro directo (Máxima prioridad)
            datos_render.get("nombre_pasajero"),
            datos_render.get("cliente_nombre"),
            datos_render.get("Cliente"),
            datos_render.get("cliente"),
            datos_render.get("pax_nombre"),
            control_int.get("cliente"),
            control_int.get("nombre_pasajero"),
            datos_render.get("titulo")
        ]
        nombre_pax = "---"
        for n in posibles_nombres:
            if n and str(n).strip() not in ["", "---", "None", "Desconocido", "Sin Título"]:
                nombre_pax = str(n).upper()
                break
        ws['B4'] = nombre_pax

        ws['D4'] = "TOTAL PAX:"
        # Mapeo robusto de PAX: PRIORIDAD PARÁMETRO > DB > CALCULADO
        p_total_param = int(num_pax or 0)
        p_total_db = int(datos_render.get("num_pasajeros") or 0)
        p_adultos = int(datos_render.get("num_adultos") or datos_render.get("adultos") or 0)
        p_ninos = int(datos_render.get("num_ninos") or datos_render.get("ninos") or 0)
        p_pax_fallback = int(datos_render.get("pax") or datos_render.get("total_pax") or 0)
        
        # Lógica de prioridad: 
        if p_total_param > 0:
            pax_count = p_total_param
        elif p_total_db > 0:
            pax_count = p_total_db
        else:
            pax_count = (p_adultos + p_ninos) if (p_adultos + p_ninos) > 0 else (p_pax_fallback if p_pax_fallback > 0 else 1)
        
        ws['E4'] = f"{pax_count} Personas"
        ws['E4'].font = bold_f

        # --- TABLA DE DÍAS ---
        current_row = 6
        headers = ["DÍA", "FECHA", "SERVICIO / TOUR", "PAX", "INCLUYE / NO INCLUYE"]
        for c_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=c_idx, value=h)
            cell.font = bold_f
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            cell.alignment = center_al
        current_row += 1

        items = (datos_render.get("itinerario_detalles") or 
                 datos_render.get("days") or 
                 datos_render.get("servicios") or 
                 datos_render.get("itinerario") or [])

        for i, d in enumerate(items):
            ws.cell(row=current_row, column=1, value=i+1).alignment = center_al
            ws.cell(row=current_row, column=2, value=d.get('fecha') or f"DIA {i+1}").alignment = center_al
            
            # Nombre del Servicio (Mapeo robusto)
            nombre_serv = (d.get('titulo') or d.get('nombre') or d.get('title') or d.get('servicio') or "---").upper()
            ws.cell(row=current_row, column=3, value=nombre_serv).font = bold_f
            
            ws.cell(row=current_row, column=4, value=pax_count).alignment = center_al
            
            # (Columna PRECIO eliminada por solicitud del usuario para reporte puramente logístico)
            # ws.cell(row=current_row, column=5, value=p_val).alignment = center_al
            
            # Detalles (Solo Inclusiones/Exclusiones, sin descripción)
            details_txt = []
            
            # Inclusiones
            inc_list = d.get('incluye') or d.get('inclusiones') or d.get('servicios') or []
            if isinstance(inc_list, list) and inc_list:
                details_txt.append("✅ INCLUYE:")
                for inc in inc_list:
                    txt = inc.get('texto') if isinstance(inc, dict) else inc
                    if txt: details_txt.append(f"  • {str(txt).upper()}")
            
            # Exclusiones
            exc_list = d.get('no_incluye') or d.get('exclusiones') or d.get('servicios_no') or []
            if isinstance(exc_list, list) and exc_list:
                if details_txt: details_txt.append("")
                details_txt.append("❌ NO INCLUYE:")
                for exc in exc_list:
                    txt = exc.get('texto') if isinstance(exc, dict) else exc
                    if txt: details_txt.append(f"  • {str(txt).upper()}")
            
            detail_cell = ws.cell(row=current_row, column=5, value="\n".join(details_txt))
            detail_cell.alignment = top_al
            detail_cell.font = Font(size=9)
            
            # Ajustar altura de fila basada en el contenido
            if details_txt:
                line_count = len(details_txt)
                ws.row_dimensions[current_row].height = max(20, line_count * 12)
            else:
                ws.row_dimensions[current_row].height = 20
            
            current_row += 1

        # --- ANCHOS DE COLUMNA ---
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 45
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 75

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def generar_hoja_servicio_maestra_xlsx(self, data_hoja: dict) -> BytesIO:
        """
        Genera un Excel Maestro Premium Consolidado en una sola pestaña.
        """
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "EXPEDIENTE_MAESTRO"
        
        # --- ESTILOS PREMIUM ---
        header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        subheader_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
        section_fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
        pax_fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
        
        white_font = Font(color="FFFFFF", bold=True)
        bold_font = Font(bold=True)
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )
        center_al = Alignment(horizontal='center', vertical='center')
        left_al = Alignment(horizontal='left', vertical='center', indent=1)

        # --- DATA PREP ---
        v = data_hoja.get('venta', {})
        it = data_hoja.get('itinerario', [])
        pax = data_hoja.get('pasajeros', [])
        liq = data_hoja.get('liquidaciones', [])

        v_moneda = v.get('moneda') or 'USD'
        v_tc_venta = float(v.get('tipo_cambio') or 3.80)
        if v_tc_venta <= 0:
            v_tc_venta = 3.80

        # --- CÁLCULO DE COSTOS MULTIMONEDA CON CONVERSIÓN ---
        costo_total_liq = 0.0
        liq_converts = []
        for l in liq:
            c_unit = float(l.get('costo_unitario') or 0)
            pax_count = int(l.get('cantidad_pax') or v.get('num_pasajeros', 1))
            total_linea = c_unit * pax_count
            
            l_moneda = l.get('moneda') or 'USD'
            tc = float(l.get('tipo_cambio') or 3.80)
            if tc <= 0:
                tc = 3.80
                
            if l_moneda == v_moneda:
                total_general = total_linea
            elif v_moneda == 'USD' and l_moneda == 'PEN':
                total_general = total_linea / tc
            elif v_moneda == 'PEN' and l_moneda == 'USD':
                total_general = total_linea * tc
            else:
                total_general = total_linea
                
            costo_total_liq += total_general
            liq_converts.append({
                'liq': l,
                'total_linea': total_linea,
                'tc': tc,
                'total_general': total_general
            })

        monto_venta = float(v.get('monto_total') or 0)
        estado_venta = v.get('estado_venta', '')
        es_cancelado = estado_venta == 'CANCELADO'
        
        ingreso_real = float(v.get('monto_pagado') or 0)
        reembolsos = float(v.get('total_reembolsado') or 0)
        
        if es_cancelado:
            utilidad = round(ingreso_real - reembolsos - costo_total_liq, 2)
            rentabilidad_str = "N/A"
        else:
            utilidad = round((monto_venta - reembolsos) - costo_total_liq, 2)
            ingreso_efectivo = monto_venta - reembolsos
            rentabilidad_str = f"{(utilidad / ingreso_efectivo * 100):.2f}%" if ingreso_efectivo > 0 else "0%"

        # --- SECCIÓN 1: PANEL DE CONTROL FINANCIERO ---
        ws.merge_cells('A1:H1')
        ws['A1'] = "REPORTE MAESTRO CONSOLIDADO: OPERACIONES & CONTABILIDAD"
        ws['A1'].font = Font(bold=True, size=16, color="1E3A8A")
        ws['A1'].alignment = center_al

        datos_v = [
            ["INFORMACIÓN GENERAL", "", ""],
            ["ID Venta", v.get('id_venta'), "Fecha Inicio"],
            ["Cliente Principal", v.get('nombre_cliente'), v.get('fecha_inicio')],
            ["Teléfono", v.get('telefono'), "Fecha Fin"],
            ["Servicio", v.get('tour_nombre'), v.get('fecha_fin')],
            ["Total Pax", f"{v.get('num_pasajeros')} PAX", "Estado Venta"],
            ["Carpeta Drive", v.get('drive_url') or "No vinculado", estado_venta or '---'],
            ["", "", ""],
            ["RESUMEN FINANCIERO", "", ""],
            ["Moneda Venta", v_moneda, "Monto Venta"],
            ["Tipo de Cambio Venta", v_tc_venta, ""],
        ]
        
        if es_cancelado:
            total_retenido = round(ingreso_real - reembolsos, 2)
            penalidad_aplicada = round(max(0.0, total_retenido - costo_total_liq), 2)
            datos_v.extend([
                ["Precio Original", monto_venta, "RECAUDADO"],
                ["Total Depositado", ingreso_real, "INGRESOS"],
                ["Total Reembolsado", reembolsos, "EGRESOS"],
                ["Total Retenido", total_retenido, "RETENIDO"],
                ["Penalidad Aplicada", penalidad_aplicada, "PENALIDAD"],
                ["Costo Operativo Real", costo_total_liq, "COSTO NETO"],
                ["BALANCE FINAL CAJA", utilidad, "RESULTADO"],
            ])
        else:
            datos_v.extend([
                ["Total Depositado", v.get('monto_pagado') or 0, "DEPOSITOS"],
                ["Total Reembolsado", v.get('total_reembolsado') or 0, "EGRESOS"],
                ["1º Método Pago", v.get('metodo_pago_primer'), f"DEP 1: {v.get('monto_primer_deposito')}"],
                ["2º Método Pago", v.get('metodo_pago_segundo'), f"DEP 2: {v.get('monto_segundo_deposito')}"],
                ["Costo Total", costo_total_liq, "COSTO NETO"],
                ["UTILIDAD ESTIMADA", utilidad, "MARGEN"],
                ["Rentabilidad", rentabilidad_str, ""],
            ])
        
        current_row = 3
        for row in datos_v:
            is_header = row[0] in ["INFORMACIÓN GENERAL", "RESUMEN FINANCIERO"]
            if is_header:
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
                cell = ws.cell(row=current_row, column=1, value=row[0])
                for c in range(1, 4):
                    ws.cell(row=current_row, column=c).fill = header_fill
                    ws.cell(row=current_row, column=c).font = white_font
                    ws.cell(row=current_row, column=c).border = thin_border
            else:
                for c_idx, val in enumerate(row, 1):
                    if val != "":
                        cell = ws.cell(row=current_row, column=c_idx, value=val)
                        cell.border = thin_border
                        # Formato especial para links
                        if row[0] == "Carpeta Drive" and c_idx == 2 and str(val).startswith("http"):
                            cell.hyperlink = val
                            cell.font = Font(color="0000FF", underline="single")
                        elif val == "CANCELADO":
                            # Resaltado premium en rojo para cancelaciones
                            cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
                            cell.font = Font(color="991B1B", bold=True)
                        elif c_idx % 2 != 0: 
                            cell.fill = subheader_fill
                            cell.font = bold_font
                        
                        # NUEVO: Formato de número de dos decimales para celdas numéricas de dinero
                        if isinstance(val, (int, float)) and row[0] not in ["ID Venta", "Tipo de Cambio Venta", "Porcentaje de Comisión Pasarela"]:
                            cell.number_format = '#,##0.00'
            current_row += 1

        current_row += 2 # Espacio

        # --- SECCIÓN 2: LOGÍSTICA DIARIA (ITINERARIO) ---
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=9)
        ws.cell(row=current_row, column=1, value="📅 CRONOGRAMA LOGÍSTICO COMPLETO").font = white_font
        for c in range(1, 10): 
            ws.cell(row=current_row, column=c).fill = section_fill
            ws.cell(row=current_row, column=c).alignment = center_al
        current_row += 1

        # Si la venta está cancelada, mostrar banner rojo visible en el cronograma
        if es_cancelado:
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=9)
            cancel_cell = ws.cell(row=current_row, column=1,
                value="⚠️  SERVICIO CANCELADO — Esta reserva fue anulada. Los servicios abajo son solo referencia histórica.")
            cancel_cell.font = Font(bold=True, color="FFFFFF", size=11)
            cancel_fill = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")
            for c in range(1, 10):
                ws.cell(row=current_row, column=c).fill = cancel_fill
                ws.cell(row=current_row, column=c).border = thin_border
            cancel_cell.alignment = center_al
            current_row += 1

        headers_it = ["Día", "Fecha", "Hora", "Servicio / Tour", "Proveedor Sugerido", "Pax", "Tipo", "Observaciones", "Estado"]
        for c_idx, h in enumerate(headers_it, 1):
            cell = ws.cell(row=current_row, column=c_idx, value=h)
            cell.fill = subheader_fill
            cell.font = bold_font
            cell.border = thin_border
            cell.alignment = center_al
        current_row += 1

        # Filtrar filas ancla autogeneradas (no deben aparecer en el reporte)
        it_filtrado = [
            s for s in it
            if not str(s.get('Servicio', '')).lower().startswith('fila auto')
            and not str(s.get('observacion', '')).lower().startswith('fila auto')
        ]

        for s in it_filtrado:
            ws.cell(row=current_row, column=1, value=s.get('Día Itin.'))
            ws.cell(row=current_row, column=2, value=s.get('Fecha'))
            ws.cell(row=current_row, column=3, value=s.get('Hora'))
            ws.cell(row=current_row, column=4, value=s.get('Servicio')).font = bold_font
            ws.cell(row=current_row, column=5, value=s.get('Proveedor'))
            ws.cell(row=current_row, column=6, value=s.get('Pax'))
            ws.cell(row=current_row, column=7, value=s.get('Tipo'))
            ws.cell(row=current_row, column=8, value=s.get('observacion') or "")
            # Estado (cancelado o vacío)
            estado_val = "CANCELADO" if es_cancelado else ""
            ws.cell(row=current_row, column=9, value=estado_val)
            # Si está cancelado, colorear toda la fila en rojo claro
            if es_cancelado:
                cancel_row_fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
                for c in range(1, 10):
                    ws.cell(row=current_row, column=c).fill = cancel_row_fill
            for c in range(1, 10): ws.cell(row=current_row, column=c).border = thin_border
            current_row += 1

        current_row += 2 # Espacio

        # --- SECCIÓN 3: LIQUIDACIÓN DE COSTOS (DETALLE PROVEEDORES) ---
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=13)
        ws.cell(row=current_row, column=1, value="💰 DETALLE DE LIQUIDACIÓN Y PAGOS").font = white_font
        for c in range(1, 14): 
            ws.cell(row=current_row, column=c).fill = section_fill
            ws.cell(row=current_row, column=c).alignment = center_al
        current_row += 1

        # Columnas actualizadas con Tipo de Cambio (T.C.) y Total General (en la moneda de venta)
        headers_l = [
            "Día", "Proveedor Real", "Tipo Servicio", "Moneda", 
            "Costo Unit.", "Cant/Pax", "Total Línea", "T.C.", 
            f"Total Gral ({v_moneda})", "Estado", "Método Pago", "Observaciones", "Obs. Contables"
        ]
        for c_idx, h in enumerate(headers_l, 1):
            cell = ws.cell(row=current_row, column=c_idx, value=h)
            cell.fill = subheader_fill
            cell.font = bold_font
            cell.border = thin_border
            cell.alignment = center_al
        current_row += 1

        red_font = Font(color="FF0000", bold=True) # Rojo para pendientes

        for item in liq_converts:
            l = item['liq']
            total_linea = item['total_linea']
            tc = item['tc']
            total_general = item['total_general']
            
            ws.cell(row=current_row, column=1, value=l.get('n_linea'))
            prov_name = l.get('proveedor', {}).get('nombre_comercial', '---') if isinstance(l.get('proveedor'), dict) else '---'
            ws.cell(row=current_row, column=2, value=prov_name).font = bold_font
            ws.cell(row=current_row, column=3, value=l.get('tipo_servicio'))
            ws.cell(row=current_row, column=4, value=l.get('moneda'))
            
            c_unit = float(l.get('costo_unitario') or 0)
            pax_count = int(l.get('cantidad_pax') or v.get('num_pasajeros', 1))
            
            ws.cell(row=current_row, column=5, value=c_unit).number_format = '#,##0.00'
            ws.cell(row=current_row, column=6, value=pax_count)
            ws.cell(row=current_row, column=7, value=total_linea).number_format = '#,##0.00'
            ws.cell(row=current_row, column=8, value=tc).number_format = '0.00'
            ws.cell(row=current_row, column=9, value=total_general).number_format = '#,##0.00'
            
            # Formateo condicional del estado
            estado_txt = "O.K." if l.get('terminado') else "Pte"
            cell_est = ws.cell(row=current_row, column=10, value=estado_txt)
            if estado_txt == "Pte":
                cell_est.font = red_font
            
            # Método de Pago, Observaciones y Obs. Contables
            ws.cell(row=current_row, column=11, value=l.get('metodo_pago', '---'))
            ws.cell(row=current_row, column=12, value=l.get('observaciones_pago', '---'))
            ws.cell(row=current_row, column=13, value=l.get('observaciones_contables', '---'))
            
            for c in range(1, 14): ws.cell(row=current_row, column=c).border = thin_border
            current_row += 1

        current_row += 2 # Espacio

        # --- SECCIÓN 4: ROOMING LIST ---
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=11)
        ws.cell(row=current_row, column=1, value="👥 LISTA DE PASAJEROS (ROOMING LIST)").font = white_font
        for c in range(1, 12): 
            ws.cell(row=current_row, column=c).fill = pax_fill
            ws.cell(row=current_row, column=c).alignment = center_al
        current_row += 1

        headers_px = ["Nombre", "Apellidos", "Documento", "Tipo Doc", "Fecha Caducidad", "Nacionalidad", "Fecha Nac", "Edad", "Género", "Cuidados", "Principal?"]
        for c_idx, h in enumerate(headers_px, 1):
            cell = ws.cell(row=current_row, column=c_idx, value=h)
            cell.fill = subheader_fill
            cell.font = bold_font
            cell.border = thin_border
        current_row += 1

        for p in pax:
            nombre_full = p.get('nombre_completo', '')
            partes = nombre_full.rsplit(' ', 1) if nombre_full else ['', '']
            ws.cell(row=current_row, column=1, value=partes[0]).font = bold_font
            ws.cell(row=current_row, column=2, value=partes[1] if len(partes) > 1 else '')
            ws.cell(row=current_row, column=3, value=p.get('numero_documento'))
            ws.cell(row=current_row, column=4, value=p.get('tipo_documento'))
            ws.cell(row=current_row, column=5, value=p.get('fecha_caducidad_doc'))
            ws.cell(row=current_row, column=6, value=p.get('nacionalidad'))
            ws.cell(row=current_row, column=7, value=p.get('fecha_nacimiento'))
            ws.cell(row=current_row, column=8, value=p.get('edad'))
            ws.cell(row=current_row, column=9, value=p.get('genero'))
            ws.cell(row=current_row, column=10, value=p.get('cuidados_especiales'))
            ws.cell(row=current_row, column=11, value="SÍ" if p.get('es_principal') else "NO")
            for c in range(1, 12): ws.cell(row=current_row, column=c).border = thin_border
            current_row += 1

        # Anchos de columna finales (Optimizados para las 12 columnas)
        col_letters = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N']
        anchos = [8, 25, 20, 10, 12, 10, 14, 10, 16, 10, 15, 30]
        for i, w in enumerate(anchos):
            ws.column_dimensions[col_letters[i]].width = w

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def generar_reporte_cuentas_cobrar_xlsx(self, df_cuentas: pd.DataFrame) -> BytesIO:
        """
        Genera un reporte Excel basado en la data de Cuentas por Cobrar B2B.
        """
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Cuentas_por_Cobrar_B2B"
        
        # --- Cabecera ---
        headers = list(df_cuentas.columns)
        for c_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c_idx, value=h)
            cell.fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid") # Azul oscuro
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # --- Datos ---
        for r_idx, row in enumerate(df_cuentas.values, 2):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        # Ajustar anchos
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 2

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def generar_ficha_control_grupos_xlsx(self, data_hoja: dict) -> BytesIO:
        """
        Genera una versión LIMPIA y ORDENADA (Pixel-Perfect) de la Ficha de Control.
        """
        import openpyxl
        from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
        from datetime import datetime
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Planilha1"
        
        # --- COLORES Y ESTILOS ---
        teal_fill = PatternFill(start_color="31859C", end_color="31859C", fill_type="solid")
        black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        white_text = Font(color="FFFFFF", bold=True, size=9) # Letra 9 para que quepa mejor
        black_bold = Font(color="000000", bold=True, size=10)
        title_f = Font(bold=True, size=22)
        detalle_font = Font(size=8)  # Letra más chica en la tabla de detalles para que no se agranden las filas
        
        border_style = Border(
            left=Side(style='thin', color='000000'), 
            right=Side(style='thin', color='000000'), 
            top=Side(style='thin', color='000000'), 
            bottom=Side(style='thin', color='000000')
        )
        
        center_al = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_al = Alignment(horizontal='left', vertical='center', indent=1)

        v = data_hoja.get('venta', {})
        pax = data_hoja.get('pasajeros', [])
        servicios = data_hoja.get('liquidaciones', [])

        # --- 0. CONFIGURACIÓN DE COLUMNAS (15 COLUMNAS: A-N usadas por el encabezado, A-O por la tabla de detalles) ---
        col_letters = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O']
        # Anchos optimizados: se ensancharon OPERADOR (F), DESCRIPCIÓN (G) y se agregó OBSERVACION (O),
        # que antes no tenía ancho definido y forzaba filas gigantes al envolver el texto.
        widths = [20, 10, 10, 10, 18, 18, 16, 10, 22, 10, 15, 18, 18, 25, 20]
        for i, w in enumerate(widths):
            ws.column_dimensions[col_letters[i]].width = w

        # --- 1. TÍTULO ---
        ws.merge_cells('A1:N1')
        ws['A1'] = "FICHA DE CONTROLE DE TOUR PARA GRUPOS"
        ws['A1'].font = title_f
        ws['A1'].alignment = center_al

        # --- 2. SECCIÓN: DATOS DEL TOUR CONDUCTOR ---
        ws.merge_cells('A2:N2')
        ws['A2'] = "DATOS DEL TOUR CONDUCTOR"
        ws['A2'].fill = black_fill
        ws['A2'].font = white_text
        ws['A2'].alignment = center_al

        # Función auxiliar para limpiar datos
        def clean(val):
            return str(val) if val and str(val).strip() not in ["", "---", "None", "0", 0] else ""

        # Preparar datos (Prioridad al Pasajero Principal del Rooming para TC)
        principal = next((p for p in pax if p.get('es_principal')), None)
        
        if principal:
            full_name = principal.get('nombre_completo', '')
            if "," in full_name:
                parts = full_name.split(",", 1)
                tc_apellido = parts[0].strip()
                tc_nombre = parts[1].strip()
            else:
                parts = full_name.split(" ", 1)
                tc_nombre = parts[0]
                tc_apellido = parts[1] if len(parts) > 1 else ""
            
            tc_nacionalidad = principal.get('nacionalidad', '')
            tc_pax = clean(principal.get('numero_documento'))
            tc_nacimiento = str(principal.get('fecha_nacimiento', ''))[:10]
            tc_caducidad = str(principal.get('fecha_caducidad_doc', ''))[:10]
        else:
            tc_nombre_full = v.get('nombre_cliente') or ""
            tc_partes = tc_nombre_full.split(' ', 1)
            tc_nombre = tc_partes[0]
            tc_apellido = tc_partes[1] if len(tc_partes) > 1 else ""
            tc_nacionalidad = ""
            tc_pax = ""
            tc_nacimiento = ""
            tc_caducidad = ""

        tc_telefono = clean(v.get('telefono_cliente') or v.get('telefono') or v.get('telefono_contacto_emergencia'))
        tc_vuelo = clean(v.get('nro_vuelo_internacional'))
        tc_correo = clean(v.get('correo_cliente'))
        tc_emer_nom = clean(v.get('nombre_contacto_emergencia'))
        tc_emer_tel = clean(v.get('telefono_contacto_emergencia'))
        tc_drive = clean(v.get('drive_url'))

        def style_label(cell):
            cell.fill = teal_fill
            cell.font = white_text
            cell.border = border_style
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False) # Sin wrap para etiquetas

        def style_value(cell):
            cell.border = border_style
            cell.alignment = center_al

        # Fila 3: NOMBRE | NACIONALIDAD | TELEFONO
        ws['A3'] = "NOMBRE"
        style_label(ws['A3'])
        ws.merge_cells('B3:D3')
        ws['B3'] = tc_nombre
        style_value(ws['B3'])
        
        ws['E3'] = "NACIONALIDAD"
        style_label(ws['E3'])
        ws.merge_cells('F3:H3')
        ws['F3'] = tc_nacionalidad
        style_value(ws['F3'])

        ws['I3'] = "TELEFONO / WHATSAPP"
        style_label(ws['I3'])
        ws.merge_cells('J3:N3')
        ws['J3'] = tc_telefono
        style_value(ws['J3'])

        # Fila 4: APELLIDO | Nº PASSAPORT | VUELO
        ws['A4'] = "APELLIDO"
        style_label(ws['A4'])
        ws.merge_cells('B4:D4')
        ws['B4'] = tc_apellido
        style_value(ws['B4'])

        ws['E4'] = "Nº PASSAPORT"
        style_label(ws['E4'])
        ws.merge_cells('F4:H4')
        ws['F4'] = tc_pax
        style_value(ws['F4'])

        ws['I4'] = "Nº VUELO INTERNACIONAL"
        style_label(ws['I4'])
        ws.merge_cells('J4:N4')
        ws['J4'] = tc_vuelo
        style_value(ws['J4'])

        # Fila 5: NASCIMENTO | CADUCIDAD | CORREO
        ws['A5'] = "FECHA DE NASCIMENTO"
        style_label(ws['A5'])
        ws.merge_cells('B5:D5')
        ws['B5'] = tc_nacimiento
        style_value(ws['B5'])

        ws['E5'] = "FECHA DE CADUCIDAD"
        style_label(ws['E5'])
        ws.merge_cells('F5:H5')
        ws['F5'] = tc_caducidad
        style_value(ws['F5'])

        ws['I5'] = "CORREO"
        style_label(ws['I5'])
        ws.merge_cells('J5:N5')
        ws['J5'] = tc_correo
        style_value(ws['J5'])

        # Fila 6: EMERGENCIA | CONTACTO | DRIVE
        ws['A6'] = "NOMBRE CONTACTO EMER"
        style_label(ws['A6'])
        ws.merge_cells('B6:D6')
        ws['B6'] = tc_emer_nom
        style_value(ws['B6'])

        ws['E6'] = "TELEFONE CONTATO EMER"
        style_label(ws['E6'])
        ws.merge_cells('F6:H6')
        ws['F6'] = tc_emer_tel
        style_value(ws['F6'])

        ws['I6'] = "LINK DEL DRIVE"
        style_label(ws['I6'])
        ws.merge_cells('J6:M6')
        ws['J6'] = tc_drive
        style_value(ws['J6'])

        # --- 3. SECCIÓN: DATOS DE VENTA ---
        ws['A7'] = "FECHA DE VENDA"
        style_label(ws['A7'])
        ws.merge_cells('B7:C7')
        ws['B7'] = clean(v.get('fecha_venta'))
        style_value(ws['B7'])
        
        ws['D7'] = "1º CIUDAD"
        style_label(ws['D7'])
        ws['E7'] = "CUSCO"
        style_value(ws['E7'])

        ws['F7'] = "FECHA DE INICIO"
        style_label(ws['F7'])
        ws.merge_cells('G7:H7')
        ws['G7'] = clean(v.get('fecha_inicio'))
        style_value(ws['G7'])

        ws['I7'] = "FECHA FINAL DEL"
        style_label(ws['I7'])
        ws.merge_cells('J7:K7')
        ws['J7'] = clean(v.get('fecha_fin'))
        style_value(ws['J7'])

        ws['L7'] = "DIAS DEL TOUR"
        style_label(ws['L7'])
        ws.merge_cells('M7:N7')
        ws['M7'] = ""
        try:
            d1 = datetime.strptime(v['fecha_inicio'], "%Y-%m-%d")
            d2 = datetime.strptime(v['fecha_fin'], "%Y-%m-%d")
            ws['M7'] = (d2 - d1).days + 1
        except: ws['M7'] = ""
        style_value(ws['M7'])

        # Fila 8: RESPONSABLE | DEPOSITO | MEDIO | TOTAL
        # Extraer agencia del origen (ej: "B2B: Kuna Travel" -> "Kuna Travel")
        origen_val = str(v.get('origen', ''))
        if origen_val.startswith('B2B:'):
            ws['A8'] = "AGENCIA / PROVEEDOR"
            origen_val = origen_val.replace('B2B:', '').strip()
        else:
            ws['A8'] = "RESPONSABLE VENDA"
            origen_val = clean(v.get('vendedor'))
            
        style_label(ws['A8'])
        ws.merge_cells('B8:D8')
        ws['B8'] = origen_val
        style_value(ws['B8'])

        ws['E8'] = "VALOR 1º DEPOSITO"
        style_label(ws['E8'])
        ws.merge_cells('F8:H8')
        ws['F8'] = clean(v.get('monto_primer_deposito'))
        style_value(ws['F8'])

        ws['I8'] = "MEDIO DE PAGO"
        style_label(ws['I8'])
        ws.merge_cells('J8:K8')
        ws['J8'] = clean(v.get('metodo_pago_primer'))
        style_value(ws['J8'])

        ws['L8'] = "TOTAL TOUR"
        style_label(ws['L8'])
        total_v = v.get('monto_total', 0)
        ws.merge_cells('M8:N8')
        ws['M8'] = f"{v.get('moneda')} {total_v}" if total_v else ""
        style_value(ws['M8'])

        # --- 4. SECCIÓN: DETALHES DEL TOUR ---
        ws.merge_cells('A9:O9')
        ws['A9'] = "DETALHES DEL TOUR"
        ws['A9'].fill = black_fill
        ws['A9'].font = white_text
        ws['A9'].alignment = center_al

        headers_tour = [
            "NOMBRE DEL TOUR", "FECHA", "HORA", "Nº DIA", "PAX", "OPERADOR", 
            "DESCRIPCIÓN DEL SERVI", "GUÍA", "MONEDA", 
            "VALOR", "PAGO", "FECHA DE CONTRATACION", "FECHA DE CONFIRMACION", "RESP. CONTRATO", "OBSERVACION"
        ]
        
        for i, h in enumerate(headers_tour, 1):
            cell = ws.cell(row=10, column=i, value=h)
            cell.fill = teal_fill
            cell.font = Font(color="FFFFFF", bold=True, size=8)
            cell.alignment = center_al
            cell.border = border_style
        
        ws.cell(row=10, column=15).border = border_style

        curr_row = 11
        itin_map = {s.get('N Linea'): s for s in data_hoja.get('itinerario', [])}

        if servicios:
            for s in servicios:
                # Buscar info extra en el itinerario usando n_linea
                nl = s.get('n_linea')
                i_info = itin_map.get(nl, {})
                
                prov_data = s.get('proveedor') or {}
                prov_nom = prov_data.get('nombre_comercial', '')
                
                f_conf = clean(s.get('fecha_confirmacion'))
                f_cont = clean(s.get('fecha_contratacion'))
                desc_serv = clean(s.get('tipo_servicio'))
                if not desc_serv: desc_serv = "SERVICIO"

                # Mapeo Exacto 1-15
                ws.cell(row=curr_row, column=1, value=clean(i_info.get('Servicio'))).border = border_style
                ws.cell(row=curr_row, column=2, value=clean(i_info.get('fecha_servicio'))).border = border_style
                ws.cell(row=curr_row, column=3, value=clean(s.get('hora_servicio'))).border = border_style
                ws.cell(row=curr_row, column=4, value=clean(i_info.get('Día Itin.'))).border = border_style
                ws.cell(row=curr_row, column=5, value=clean(s.get('cantidad_pax'))).border = border_style
                ws.cell(row=curr_row, column=6, value=clean(prov_nom)).border = border_style
                ws.cell(row=curr_row, column=7, value=desc_serv).border = border_style
                ws.cell(row=curr_row, column=8, value=clean(s.get('nombre_guia'))).border = border_style
                ws.cell(row=curr_row, column=9, value=clean(s.get('moneda'))).border = border_style
                ws.cell(row=curr_row, column=10, value=clean(s.get('costo_unitario'))).border = border_style
                ws.cell(row=curr_row, column=11, value="").border = border_style # Pago
                ws.cell(row=curr_row, column=12, value=f_cont).border = border_style
                ws.cell(row=curr_row, column=13, value=f_conf).border = border_style
                ws.cell(row=curr_row, column=14, value=clean(s.get('responsable_contratacion'))).border = border_style
                ws.cell(row=curr_row, column=15, value=clean(s.get('observacion'))).border = border_style

                for c in range(1, 16):
                    cell_det = ws.cell(row=curr_row, column=c)
                    cell_det.alignment = center_al
                    cell_det.font = detalle_font
                ws.row_dimensions[curr_row].height = 28  # Altura fija para que el ajuste de texto no dispare filas gigantes
                curr_row += 1
        
        # Filas extra para completar el diseño (total 15 filas de tour)
        while curr_row < 25:
            for c in range(1, 16):
                ws.cell(row=curr_row, column=c).border = border_style
            ws.row_dimensions[curr_row].height = 28
            curr_row += 1

        # --- 5. PASAJEROS / ROOMING LIST ---
        curr_row += 1
        ws.merge_cells(f'A{curr_row}:N{curr_row}')
        ws[f'A{curr_row}'] = "PASAJEROS / ROOMING LIST"
        ws[f'A{curr_row}'].fill = black_fill
        ws[f'A{curr_row}'].font = white_text
        ws[f'A{curr_row}'].alignment = center_al
        curr_row += 1

        pax_config = [
            ("Nº", 1, 1), ("APELLIDOS / NOMBRES", 2, 5), ("SEXO", 6, 6),
            ("PASAPORTE", 7, 8), ("NACIONALIDAD", 9, 9), ("FECHA NAC.", 10, 10),
            ("CUIDADOS", 11, 12), ("HABITACIÓN", 13, 14)
        ]

        for h, s_c, e_c in pax_config:
            cell = ws.cell(row=curr_row, column=s_c, value=h)
            cell.fill = teal_fill
            cell.font = white_text
            cell.border = border_style
            cell.alignment = center_al
            if s_c != e_c: ws.merge_cells(start_row=curr_row, end_row=curr_row, start_column=s_c, end_column=e_c)
            for c in range(s_c, e_c+1): ws.cell(row=curr_row, column=c).border = border_style
        
        curr_row += 1
        for i, p in enumerate(pax, 1):
            ws.cell(row=curr_row, column=1, value=i).border = border_style
            ws.cell(row=curr_row, column=1).alignment = center_al
            
            ws.cell(row=curr_row, column=2, value=str(p.get('nombre_completo', '')).upper()).border = border_style
            ws.merge_cells(start_row=curr_row, end_row=curr_row, start_column=2, end_column=5)
            
            ws.cell(row=curr_row, column=6, value=clean(p.get('genero'))).border = border_style
            ws.cell(row=curr_row, column=6).alignment = center_al
            
            ws.cell(row=curr_row, column=7, value=clean(p.get('numero_documento'))).border = border_style
            ws.merge_cells(start_row=curr_row, end_row=curr_row, start_column=7, end_column=8)
            ws.cell(row=curr_row, column=7).alignment = center_al
            
            ws.cell(row=curr_row, column=9, value=clean(p.get('nacionalidad'))).border = border_style
            ws.cell(row=curr_row, column=10, value=str(p.get('fecha_nacimiento', ''))[:10]).border = border_style
            
            ws.cell(row=curr_row, column=11, value=clean(p.get('cuidados_especiales'))).border = border_style
            ws.merge_cells(start_row=curr_row, end_row=curr_row, start_column=11, end_column=12)
            
            ws.cell(row=curr_row, column=13, value=clean(p.get('acomodacion'))).border = border_style
            ws.merge_cells(start_row=curr_row, end_row=curr_row, start_column=13, end_column=14)
            
            for c in range(1, 15): 
                ws.cell(row=curr_row, column=c).border = border_style
                ws.cell(row=curr_row, column=c).alignment = center_al
            curr_row += 1

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def generar_reporte_ventas_consolidado_xlsx(self, supabase_client) -> BytesIO:
        """Genera el Reporte Maestro de Ventas (Tabla Consolidada)."""
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        import datetime
        from io import BytesIO

        try:
            # Traer ventas, datos del cliente/lead, pasajeros y pagos
            res_ventas = supabase_client.table('venta').select('*, cliente(nombre, lead(pais_origen)), pasajero(nacionalidad, edad, genero, es_principal)').order('fecha_venta', desc=True).execute()
            ventas_data = res_ventas.data or []
            
            res_pagos = supabase_client.table('pago').select('*').order('fecha_pago', desc=False).execute()
            pagos_data = res_pagos.data or []
        except Exception as e:
            print(f"Error fetching data for master report: {e}")
            ventas_data = []
            pagos_data = []

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "REPORTE MAESTRO VENTAS"

        # --- ESTILOS ---
        header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid") # Azul
        white_text = Font(color="FFFFFF", bold=True)
        center_al = Alignment(horizontal='center', vertical='center', wrap_text=True)
        border_style = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        headers = [
            "FECHA VENTA", "PAX", "NACIONALIDAD", "EDAD", "GÉNERO", "FECHA DE TOUR", 
            "MONTO TOTAL S/.", "MONTO TOTAL $", 
            "PRIMER PAGO (FECHA)", "PRIMER PAGO S/.", "PRIMER PAGO $", "PRIMER PAGO MODALIDAD",
            "SEGUNDO PAGO (FECHA)", "SEGUNDO PAGO S/.", "SEGUNDO PAGO $", "SEGUNDO PAGO MODALIDAD",
            "TOTAL PAGADO S/.", "TOTAL PAGADO $", 
            "DIFERENCIA S/.", "DIFERENCIA $", 
            "ESTADO", "TIPO (B2C/B2B)"
        ]

        for col_num, header_title in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header_title)
            cell.fill = header_fill
            cell.font = white_text
            cell.alignment = center_al
            cell.border = border_style
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 18

        # Agrupar pagos por venta
        pagos_por_venta = {}
        for p in pagos_data:
            id_v = p.get('id_venta')
            if id_v not in pagos_por_venta:
                pagos_por_venta[id_v] = []
            pagos_por_venta[id_v].append(p)

        curr_row = 2
        for v in ventas_data:
            cliente_info = v.get('cliente') or {}
            lead_info = cliente_info.get('lead') or {} if isinstance(cliente_info, dict) else {}
            pax_info = v.get('pasajero', [])
            
            # Fechas y Básicos
            f_venta = v.get('fecha_venta')
            nombre_pax = v.get('nombre_cliente') or cliente_info.get('nombre', 'Desconocido')
            pax = f"{nombre_pax} x {v.get('num_pasajeros', 1)}"
            
            # Lógica para Nacionalidad, Edad y Género (Pasajero Principal > Lead > Blanco)
            nacionalidad = ""
            edad = ""
            genero = ""
            if pax_info:
                principal = next((p for p in pax_info if p.get('es_principal')), pax_info[0])
                nacionalidad = principal.get('nacionalidad') or ""
                edad = principal.get('edad') or ""
                genero = principal.get('genero') or ""
            else:
                nacionalidad = lead_info.get('pais_origen') or ""
                
            # Limpiar la palabra 'Nacional' si viene de la base de datos
            if isinstance(nacionalidad, str) and nacionalidad.lower() == 'nacional':
                nacionalidad = ""
                
            f_tour = v.get('fecha_inicio', '')
            
            # Montos
            moneda_venta = v.get('moneda', 'USD')
            monto_orig = float(v.get('precio_total_cierre') or 0.0)
            tc = float(v.get('tipo_cambio') or 3.80)
            
            if moneda_venta == 'USD':
                monto_total_usd = monto_orig
                monto_total_pen = monto_orig * tc
            else:
                monto_total_pen = monto_orig
                monto_total_usd = monto_orig / tc if tc > 0 else 0.0

            # Pagos
            lista_pagos = pagos_por_venta.get(v.get('id_venta'), [])
            # Filtrar reembolsos
            lista_pagos = [p for p in lista_pagos if p.get('tipo_pago') != 'REEMBOLSO']
            
            p1_fecha, p1_pen, p1_usd, p1_mod = "", 0.0, 0.0, ""
            p2_fecha, p2_pen, p2_usd, p2_mod = "", 0.0, 0.0, ""
            total_pagado_pen, total_pagado_usd = 0.0, 0.0

            if len(lista_pagos) > 0:
                p1 = lista_pagos[0]
                p1_fecha = p1.get('fecha_pago', '')
                p1_mod = p1.get('metodo_pago', '')
                p_moneda = p1.get('moneda', moneda_venta)
                p_monto = float(p1.get('monto_pagado') or 0.0)
                if p_moneda == 'USD':
                    p1_usd = p_monto
                    p1_pen = p_monto * tc
                else:
                    p1_pen = p_monto
                    p1_usd = p_monto / tc if tc > 0 else 0.0

            if len(lista_pagos) > 1:
                p2 = lista_pagos[1]
                p2_fecha = p2.get('fecha_pago', '')
                p2_mod = p2.get('metodo_pago', '')
                p_moneda = p2.get('moneda', moneda_venta)
                p_monto = float(p2.get('monto_pagado') or 0.0)
                if p_moneda == 'USD':
                    p2_usd = p_monto
                    p2_pen = p_monto * tc
                else:
                    p2_pen = p_monto
                    p2_usd = p_monto / tc if tc > 0 else 0.0
            
            for p in lista_pagos:
                p_moneda = p.get('moneda', moneda_venta)
                p_monto = float(p.get('monto_pagado') or 0.0)
                if p_moneda == 'USD':
                    total_pagado_usd += p_monto
                    total_pagado_pen += p_monto * tc
                else:
                    total_pagado_pen += p_monto
                    total_pagado_usd += p_monto / tc if tc > 0 else 0.0
            
            dif_pen = monto_total_pen - total_pagado_pen
            dif_usd = monto_total_usd - total_pagado_usd
            
            estado = str(v.get('estado_venta', 'CONFIRMADO')).upper()
            tipo = "B2B" if v.get('id_agencia_aliada') else "B2C"

            row_data = [
                f_venta, pax, nacionalidad, edad, genero, f_tour,
                monto_total_pen, monto_total_usd,
                p1_fecha, p1_pen, p1_usd, p1_mod,
                p2_fecha, p2_pen, p2_usd, p2_mod,
                total_pagado_pen, total_pagado_usd,
                dif_pen, dif_usd,
                estado, tipo
            ]
            
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=curr_row, column=col_num, value=value)
                cell.border = border_style
                cell.alignment = center_al
                # Format numbers
                if isinstance(value, float):
                    cell.number_format = '#,##0.00'
            
            if estado == 'CANCELADO':
                for col_num in range(1, len(row_data) + 1):
                    ws.cell(row=curr_row, column=col_num).fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
                    ws.cell(row=curr_row, column=col_num).font = Font(color="991B1B")

            curr_row += 1

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output


